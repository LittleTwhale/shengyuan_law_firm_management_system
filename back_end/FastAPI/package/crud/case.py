# crud/case.py
import json
import os
from datetime import datetime
from io import BytesIO
from typing import List
from typing import Optional, cast

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from sqlalchemy import func, and_
from sqlalchemy import or_, extract
from sqlalchemy.orm import Session, joinedload, selectinload

from ..models.case import Case, CaseParty, BankCase
from ..models.electronic_volume_model import CaseVolume, VolumeFile
from ..models.finance_model import CaseFinance
from ..models.attachment import CaseAttachment
from ..models.user import User
from ..models.user import UserSchedule
from ..schemas.case import CaseCreate, CaseUpdate
from ..schemas.case import CaseExportQuery
from ..core.config import CASE_ATTACHMENT_ROOT, settings


def get_case_by_id(db: Session, case_id: int) -> Optional[Case]:
    """
    根据案件ID获取案件
    """
    return (
        db.query(Case)
        .options(
            joinedload(Case.main_lawyer),
            joinedload(Case.assistant_lawyer),
            joinedload(Case.assistant_lawyer_2),
            joinedload(Case.execution_lawyer),
            joinedload(Case.execution_assistant),
            joinedload(Case.bank_case_details),
            joinedload(Case.parties),
        )
        .filter(
            Case.case_id == case_id,
            Case.is_deleted == False
        )
        .first()
    )


def can_user_view_case(user: User, case: Case) -> bool:
    """
    判断用户是否有权查看案件详情（与列表接口的过滤规则保持一致）

    规则：
    - admin / owner：可查看全部案件
    - 普通用户：仅可查看自己参与的案件（主办/助理/第二助理/执行主办/执行助理）
    - 普通用户拥有 can_view_all_bank_events 权限时：额外可查看全部银行案件
    """
    if user.role in ["admin", "owner"]:
        return True

    user_id = user.id
    if user_id in (
        case.main_lawyer_id,
        case.assistant_lawyer_id,
        case.assistant_lawyer_2_id,
        case.execution_lawyer_id,
        case.execution_assistant_id,
    ):
        return True

    perms = user.permissions or {}
    if perms.get("can_view_all_bank_events") and case.case_category == "银行案件":
        return True

    return False


def can_user_edit_case(user: User, case: Case) -> bool:
    """
    判断用户是否有权修改/删除案件（写权限，比查看权限更严格）

    规则：
    - admin / owner：可修改/删除全部案件
    - 普通用户：仅可修改/删除自己参与的案件（主办/助理/第二助理/执行主办/执行助理）
    - 注意：can_view_all_bank_events 仅是"查看"权限，不授予修改他人案件的权力
    """
    if user.role in ["admin", "owner"]:
        return True

    user_id = user.id
    if user_id in (
        case.main_lawyer_id,
        case.assistant_lawyer_id,
        case.assistant_lawyer_2_id,
        case.execution_lawyer_id,
        case.execution_assistant_id,
    ):
        return True

    return False


def list_cases_by_user_role(
        db: Session,
        user_id: int,
        role: str,
        skip: int = 0,
        limit: int = 100,
        keyword: Optional[str] = None,  # 关键词查询
        category: Optional[str] = None,  # 案件类型筛选
        main_lawyer_id: Optional[int] = None,  # 主办律师筛选
        execution_lawyer_id: Optional[int] = None, # 执行主办律师筛选
        year: Optional[str] = None,  # 年份筛选
        review_status: Optional[str] = None,  # 审核状态筛选
        sort_field: str = "created_at",  # 排序字段，默认按创建时间
        sort_dir: str = "desc",  # 排序方向，默认降序（最新在前）
        can_view_all_bank: bool = False # 是否允许查看所有银行案件
) -> List[Case]:
    """
    根据用户角色返回案件列表
    - 普通用户：只能看到自己为主办律师或协办律师的案件
    - admin/owner/can_view_all_bank_events：可以看到全部案件
    """
    query = db.query(Case).options(
        joinedload(Case.main_lawyer),
        joinedload(Case.assistant_lawyer),
        joinedload(Case.assistant_lawyer_2),
        joinedload(Case.execution_lawyer),
        joinedload(Case.execution_assistant),
        joinedload(Case.parties)
    ).filter(Case.is_deleted == False)

    if year:
        # 使用 extract 提取委托日期的年份进行匹配
        query = query.filter(Case.case_number.like(f"%({year})%"))

    # 角色与主办律师筛选逻辑
    if role not in ["admin", "owner"]:
        user_involved_cond = or_(
            Case.main_lawyer_id == user_id,
            Case.assistant_lawyer_id == user_id,
            Case.assistant_lawyer_2_id == user_id,
            Case.execution_lawyer_id == user_id,
            Case.execution_assistant_id == user_id
        )
        if can_view_all_bank:
            # 有权限：能看自己的所有案件 + 别人的银行案件
            query = query.filter(or_(user_involved_cond, Case.case_category == "银行案件"))
        else:
            # 无权限：只能看自己相关的案件
            query = query.filter(user_involved_cond)

    if main_lawyer_id is not None:
        query = query.filter(Case.main_lawyer_id == main_lawyer_id)

    # 执行主办律师过滤
    if execution_lawyer_id is not None:
        query = query.filter(Case.execution_lawyer_id == execution_lawyer_id)

    # 类别筛选
    if category:
        query = query.filter(Case.case_category == category)

    # 关键词搜索：支持按案号或【任何当事人名称】进行全维度检索
    if keyword:
        query = query.filter(
            or_(
                Case.case_number.like(f"%{keyword}%"),
                Case.parties.any(CaseParty.name.like(f"%{keyword}%"))
            )
        )

    # 审核状态筛选
    if review_status:
        query = query.filter(Case.review_status == review_status)

    # 排序逻辑
    if sort_field == "created_at":
        order_column = Case.created_at
    elif sort_field == "updated_at":
        order_column = Case.updated_at
    elif sort_field == "commission_date":  # 委托日期
        order_column = Case.commission_date
    else:
        order_column = Case.created_at  # 默认字段

    if sort_dir == "desc":
        query = query.order_by(order_column.desc())
    else:
        query = query.order_by(order_column.asc())

    cases = query.offset(skip).limit(limit).all()
    return cast(list[Case], cast(object, cases))


def count_cases_by_user_role(
        db: Session,
        user_id: int,
        role: str,
        keyword: Optional[str] = None,  # 关键词查询
        category: Optional[str] = None,  # 案件类型筛选
        main_lawyer_id: Optional[int] = None,  # 主办律师筛选
        execution_lawyer_id: Optional[int] = None, # 执行主办律师筛选
        year: Optional[str] = None,  # 年份筛选
        review_status: Optional[str] = None,  # 审核状态筛选
        can_view_all_bank: bool = False # 是否允许查看所有银行案件
) -> int:
    """
    根据用户角色统计案件总数
    """
    query = db.query(Case).filter(Case.is_deleted == False)
    if year:
        query = query.filter(Case.case_number.like(f"%({year})%"))

    # 角色筛选
    if role not in ["admin", "owner"]:
        user_involved_cond = or_(
            Case.main_lawyer_id == user_id,
            Case.assistant_lawyer_id == user_id,
            Case.assistant_lawyer_2_id == user_id,
            Case.execution_lawyer_id == user_id,
            Case.execution_assistant_id == user_id
        )
        if can_view_all_bank:
            # 有权限：能看自己的所有案件 + 别人的银行案件
            query = query.filter(or_(user_involved_cond, Case.case_category == "银行案件"))
        else:
            # 无权限：只能看自己相关的案件
            query = query.filter(user_involved_cond)

    if main_lawyer_id is not None:
        query = query.filter(Case.main_lawyer_id == main_lawyer_id)

    if execution_lawyer_id is not None:
        query = query.filter(Case.execution_lawyer_id == execution_lawyer_id)

    # 类别筛选
    if category:
        query = query.filter(Case.case_category == category)

    # 关键词搜索：支持按案号或【任何当事人名称】进行全维度检索
    if keyword:
        query = query.filter(
            or_(
                Case.case_number.like(f"%{keyword}%"),
                Case.parties.any(CaseParty.name.like(f"%{keyword}%"))
            )
        )

    # 审核状态筛选
    if review_status:
        query = query.filter(Case.review_status == review_status)

    return query.count()


def list_bank_cases_by_user_role(
        db: Session,
        user_id: int,
        role: str,
        skip: int = 0,
        limit: int = 100,
        keyword: Optional[str] = None,
        main_lawyer_id: Optional[int] = None,  # 主办律师筛选
        execution_lawyer_id: Optional[int] = None, # 执行主办律师筛选
        client_name: Optional[str] = None, # 委托银行筛选
        year: Optional[str] = None,
        case_status: Optional[str] = None,
        sort_field: str = "created_at",  # 排序字段，默认按创建时间
        sort_dir: str = "desc",  # 排序方向，默认降序（最新在前）
        can_view_all_bank: bool = False # 是否允许查看所有银行案件
) -> List[Case]:
    """
    根据用户角色返回银行案件列表
    - 普通用户：只能看到自己为主办律师或协办的案件
    - admin/owner：可以看到全部案件
    """
    query = db.query(Case).options(
        joinedload(Case.main_lawyer),
        joinedload(Case.assistant_lawyer),
        joinedload(Case.assistant_lawyer_2),
        joinedload(Case.execution_lawyer),
        joinedload(Case.execution_assistant),
        joinedload(Case.parties),
        joinedload(Case.bank_case_details)
    ).filter(Case.is_deleted == False, Case.case_category == "银行案件")

    # 角色筛选
    if role not in ["admin", "owner"] and not can_view_all_bank:
        query = query.filter(
            or_(
                Case.main_lawyer_id == user_id,
                Case.assistant_lawyer_id == user_id,
                Case.assistant_lawyer_2_id == user_id,
                Case.execution_lawyer_id == user_id,
                Case.execution_assistant_id == user_id
            )
        )

    if main_lawyer_id is not None:
        query = query.filter(Case.main_lawyer_id == main_lawyer_id)

    if execution_lawyer_id is not None:
        query = query.filter(Case.execution_lawyer_id == execution_lawyer_id)

    # 委托银行筛选逻辑 (通过 CaseParty 表查询)
    if client_name:
        query = query.filter(
            Case.parties.any(
                and_(CaseParty.party_type.like('%委托%'), CaseParty.name.like(f"%{client_name}%"))
            )
        )

    # 关键词搜索：支持按案号、法院案号或【任何当事人名称】进行全维度检索
    if keyword:
        query = query.filter(
            or_(
                Case.case_number.like(f"%{keyword}%"),
                Case.case_code.like(f"%{keyword}%"),
                Case.parties.any(CaseParty.name.like(f"%{keyword}%"))
            )
        )

    # 委托年份筛选
    if year:
        query = query.filter(Case.case_number.like(f"%({year})%"))

    # 案件状态筛选
    if case_status:
        # join BankCase 表进行筛选
        query = query.join(BankCase).filter(BankCase.case_status == case_status)

    # 排序逻辑
    if sort_field == "created_at":
        order_column = Case.created_at
    elif sort_field == "updated_at":
        order_column = Case.updated_at
    elif sort_field == "commission_date":  # 委托日期
        order_column = Case.commission_date
    else:
        order_column = Case.created_at  # 默认字段

    if sort_dir == "desc":
        query = query.order_by(order_column.desc())
    else:
        query = query.order_by(order_column.asc())

    cases = query.offset(skip).limit(limit).all()
    return cast(list[Case], cast(object, cases))


def count_bank_cases_by_user_role(
        db: Session,
        user_id: int,
        role: str,
        keyword: Optional[str] = None,
        main_lawyer_id: Optional[int] = None,
        execution_lawyer_id: Optional[int] = None,
        client_name: Optional[str] = None,
        year: Optional[str] = None,
        case_status: Optional[str] = None,
        can_view_all_bank: bool = False
) -> int:
    """
    根据用户角色统计案件总数
    """
    query = db.query(Case).filter(Case.is_deleted == False, Case.case_category == "银行案件")
    # 角色筛选
    if role not in ["admin", "owner"] and not can_view_all_bank:
        query = query.filter(
            or_(
                Case.main_lawyer_id == user_id,
                Case.assistant_lawyer_id == user_id,
                Case.assistant_lawyer_2_id == user_id,
                Case.execution_lawyer_id == user_id,
                Case.execution_assistant_id == user_id
            )
        )

    if main_lawyer_id is not None:
        query = query.filter(Case.main_lawyer_id == main_lawyer_id)

    if execution_lawyer_id is not None:
        query = query.filter(Case.execution_lawyer_id == execution_lawyer_id)

    # 委托银行筛选逻辑 (通过 CaseParty 表查询)
    if client_name:
        query = query.filter(
            Case.parties.any(
                and_(CaseParty.party_type.like('%委托%'), CaseParty.name.like(f"%{client_name}%"))
            )
        )

    # 关键词搜索：支持按案号、法院案号或【任何当事人名称】进行全维度检索
    if keyword:
        query = query.filter(
            or_(
                Case.case_number.like(f"%{keyword}%"),
                Case.case_code.like(f"%{keyword}%"),
                Case.parties.any(CaseParty.name.like(f"%{keyword}%"))
            )
        )

    # 委托年份筛选
    if year:
        query = query.filter(Case.case_number.like(f"%({year})%"))

    # 案件状态筛选
    if case_status:
        query = query.join(BankCase).filter(BankCase.case_status == case_status)

    return query.count()


def create_case(db: Session, case_in: CaseCreate) -> Case:
    """
    创建新案件（复用已删除案件的原始编号，但创建新记录）
    """
    year = datetime.now().year

    # 案件类型映射
    type_map = {
        "民事案件": "民字",
        "刑事案件": "刑字",
        "劳动仲裁": "劳仲字",
        "商事仲裁": "商仲字",
        "行政案件": "行字",
        "非诉业务": "非诉字",
        "执行案件": "执行字",
        "法律顾问业务": "法顾字",
        "银行案件": "银行案件",
        "法律援助(民事)": "法律援助(民)",
        "法律援助(刑事)": "法律援助(刑)",
        "法律援助(行政)": "法律援助(行)",
    }

    case_type = case_in.case_category
    if case_type not in type_map:
        raise ValueError("未知的案件类型")

    if case_in.case_number:
        # 校验手动传入的案件号是否已存在，防止触发 models 中 case_number unique=True 的报错
        existing_case = db.query(Case).filter(Case.case_number == case_in.case_number).first()
        if existing_case:
            raise ValueError(f"案件号 {case_in.case_number} 已存在，请检查导入表格")
        final_case_number = case_in.case_number
    else:
        # 如果没有传入案件号，执行原有的自动生成逻辑
        final_case_number = _find_reusable_case_number(db, case_type, year)

    # 清洗当事人列表空字段
    if case_in.parties:
        for party in case_in.parties:
            for key, value in party.model_dump().items():
                if isinstance(value, str) and value.strip() in ("None", "nan", "NaN", "", "null"):
                    # 使用 setattr 修改 Pydantic 对象的属性，而不是 party[key]
                    setattr(party, key, None)

    # 创建全新的案件记录，但使用复用的编号
    # 分离 Case 数据和 BankCase、parties 数据
    case_data = case_in.model_dump(exclude={"bank_case_details", "parties", "case_number"})

    # 清洗数据，将字面量字符串 "None", "nan", "" 转换为真正的 None
    for key, value in case_data.items():
        if isinstance(value, str):
            # 去除首尾空格后，如果是这些无效值，就转为 None
            if value.strip() in ("None", "nan", "NaN", "", "null"):
                case_data[key] = None

    # 创建主案件
    case_data["review_status"] = "待审核"
    case_data["is_deleted"] = False
    new_case = Case(**case_data, case_number=final_case_number)
    db.add(new_case)
    db.flush()  # 刷新以获取 new_case.case_id

    # 案件创建时，自动建立一对一的财务关联
    initial_amount = new_case.case_income or 0
    new_finance = CaseFinance(
        case_id=new_case.case_id,
        contract_amount=initial_amount,
        unpaid_amount=initial_amount,  # 初始欠款直接等于合同金额
        uninvoiced_amount=initial_amount,  # 初始未开票直接等于合同金额
        remarks="案件创建自动初始化"
    )
    db.add(new_finance)

    # 保存当事人列表
    if case_in.parties:
        for party in case_in.parties:
            new_party = CaseParty(
                case_id=new_case.case_id,
                **party.model_dump()
            )
            db.add(new_party)

    # 如果是银行案件且提供了详情，则创建扩展表记录
    if case_in.case_category == "银行案件" and case_in.bank_case_details:
        bank_data = case_in.bank_case_details.model_dump()
        # 清洗银行案件中的字符串 "None"
        for key, value in bank_data.items():
            if isinstance(value, str) and value.strip() in ("None", "nan", "NaN", "", "null"):
                bank_data[key] = None
        new_bank_case = BankCase(case_id=new_case.case_id, **bank_data)
        db.add(new_bank_case)

    db.commit()
    db.refresh(new_case)
    return new_case


def _find_reusable_case_number(db: Session, case_type: str, year: int) -> str:
    """
    查找可复用的案件编号
    """
    type_map = {
        "民事案件": "民字",
        "刑事案件": "刑字",
        "劳动仲裁": "劳仲字",
        "商事仲裁": "商仲字",
        "行政案件": "行字",
        "非诉业务": "非诉字",
        "执行案件": "执行字",
        "法律顾问业务": "法顾字",
        "银行案件": "银行案件",
        "法律援助(民事)": "法律援助(民)",
        "法律援助(刑事)": "法律援助(刑)",
        "法律援助(行政)": "法律援助(行)",
    }

    # 查找已删除的案件，按案件号排序（找到最小的可用编号）
    deleted_cases = db.query(Case).filter(
        Case.case_category == case_type,
        Case.is_deleted == True,
        Case.case_number.like(f"[已删除]湘生律({year})%")
    ).order_by(Case.case_number).all()

    # 优先复用已删除案件的原始编号
    for deleted_case in deleted_cases:
        original_number = deleted_case.case_number.replace("[已删除]", "").split("-ID")[0]

        # 检查这个原始编号是否已被其他活跃案件使用
        existing_active_case = db.query(Case).filter(
            Case.case_number == original_number,
            Case.is_deleted == False
        ).first()

        if not existing_active_case:
            # 这个编号可用，直接返回
            return original_number

    # 如果没有可复用的编号，创建新编号
    return _create_new_case_number(db, case_type, year, type_map)


def _create_new_case_number(db: Session, case_type: str, year: int, type_map: dict) -> str:
    """
    创建新案件编号（修复版：解决ID顺序与案号顺序不一致导致的冲突）
    """
    # 尝试基于最新ID的案件推算（但这不一定是最大案号）
    latest_case = db.query(Case).filter(
        Case.case_category == case_type,
        Case.case_number.like(f"湘生律({year})%")
        # 去掉 is_deleted 限制，确保基于所有历史数据递增
    ).order_by(Case.case_id.desc()).first()

    next_number = 1
    if latest_case:
        try:
            # 提取数字
            if "第" in latest_case.case_number and "号" in latest_case.case_number:
                last_number = int(latest_case.case_number.split("第")[-1].replace("号", ""))
                next_number = last_number + 1
        except (ValueError, IndexError):
            next_number = 1

    # 循环检测，直到找到一个真正空闲的号码
    while True:
        candidate = f"湘生律({year}){type_map[case_type]}第{next_number}号"

        # 检查候选号码是否已存在（包含已删除但未改名的）
        exists = db.query(Case).filter(Case.case_number == candidate).first()

        if not exists:
            return candidate

        # 如果存在，递增序号并重试
        next_number += 1


def update_case(db: Session, case_id: int, case_in: CaseUpdate) -> Optional[Case]:
    """
    更新已有案件
    """
    case = db.query(Case).options(
        joinedload(Case.main_lawyer),
        joinedload(Case.assistant_lawyer),
        joinedload(Case.assistant_lawyer_2),
        joinedload(Case.execution_lawyer),
        joinedload(Case.execution_assistant),
        joinedload(Case.bank_case_details),
    ).filter(
        Case.case_id == case_id,
        Case.is_deleted == False
    ).first()

    if not case:
        return None

    # 保存旧类型，提取新类型（注意可能未传入新类型）
    old_category = case.case_category
    new_category = case_in.case_category or old_category

    # 检查是否修改了案件类别
    category_changed = (new_category != old_category)

    # 更新主表数据
    # 先批量更新其他字段，但跳过案件类别，避免提前改变
    case_data = case_in.model_dump(exclude_unset=True, exclude={"bank_case_details", "parties"})
    for key, value in case_data.items():
        if key == "case_category": continue
        setattr(case, key, value)

    # 处理当事人列表更新 (全删全增策略)
    if case_in.parties is not None:
        # A. 删除该案件所有的旧当事人
        db.query(CaseParty).filter(CaseParty.case_id == case_id).delete()

        # B. 添加新当事人
        for party in case_in.parties:
            new_party = CaseParty(
                case_id=case_id,
                **party.model_dump()
            )
            db.add(new_party)

    # 更新或创建银行案件详情
    if case_in.bank_case_details:
        if case.bank_case_details:
            # 更新现有记录
            bank_update_data = case_in.bank_case_details.model_dump(exclude_unset=True)
            for k, v in bank_update_data.items():
                setattr(case.bank_case_details, k, v)
        else:
            # 如果之前没有详情（可能是从其他类型转过来的），则创建
            if case.case_category == "银行案件":
                bank_data = case_in.bank_case_details.model_dump()
                new_bank_case = BankCase(case_id=case.case_id, **bank_data)
                db.add(new_bank_case)

    # 如果案件类别发生变化，则重新生成编号
    if category_changed:
        year = datetime.now().year
        type_map = {
            "民事案件": "民字",
            "刑事案件": "刑字",
            "劳动仲裁": "劳仲字",
            "商事仲裁": "商仲字",
            "行政案件": "行字",
            "非诉业务": "非诉字",
            "执行案件": "执行字",
            "法律顾问业务": "法顾字",
            "银行案件": "银行案件",
            "法律援助(民事)": "法律援助(民)",
            "法律援助(刑事)": "法律援助(刑)",
            "法律援助(行政)": "法律援助(行)",
        }

        if new_category not in type_map:
            raise ValueError("未知的案件类型")

        # 查询该类型最新案件号
        latest_case = db.query(Case).filter(
            Case.case_category == new_category,
            Case.case_number.like(f"湘生律({year})%")
        ).order_by(Case.case_id.desc()).first()

        next_number = 1
        if latest_case:
            try:
                last_number = int(latest_case.case_number.split("第")[-1].replace("号", ""))
                next_number = last_number + 1
            except ValueError:
                next_number = 1  # 容错处理

        # 生成新编号并更新类型
        case.case_category = new_category
        case.case_number = f"湘生律({year}){type_map[new_category]}第{next_number}号"


    # 如果案件是审核不通过或业务类型发生变更，统一设置为待审核,并将审核人设为空
    if case.review_status == "已拒绝" or category_changed:
        case.review_status = "待审核"
        case.reviewer_id = None
        case.reviewed_at = None
        case.review_comment = None  # 清空旧的审核意见

    db.commit()
    db.refresh(case)
    return cast(Case, case)


def delete_case(db: Session, case_id: int) -> bool:
    """
    删除案件（逻辑删除 Case，但在物理上清理关联的卷宗文件和财务数据）
    """
    case = db.query(Case).filter(Case.case_id == case_id, Case.is_deleted == False).first()
    if not case:
        return False

    # =========================================================
    # 1. 清理电子卷宗 (数据库记录 + 物理文件)
    # =========================================================
    # 查询该案件下的所有卷册
    volumes = db.query(CaseVolume).filter(CaseVolume.case_id == case_id).all()

    for volume in volumes:
        # A. 删除卷册合并的 PDF (如果有)
        if volume.merged_file_path and os.path.exists(volume.merged_file_path):
            try:
                os.remove(volume.merged_file_path)
            except OSError as e:
                print(f"删除合并卷宗文件失败: {e}")

        # B. 删除卷内文件
        # 注意：这里需要查询 VolumeFile，因为 CaseVolume 和 VolumeFile 配置了 cascade，
        # 直接 db.delete(volume) 会删数据库记录，但不会删磁盘文件，所以我们要手动遍历。
        files = db.query(VolumeFile).filter(VolumeFile.volume_id == volume.id).all()
        for file_obj in files:
            if file_obj.file_path and os.path.exists(file_obj.file_path):
                try:
                    os.remove(file_obj.file_path)
                except OSError as e:
                    print(f"删除物理文件失败: {file_obj.file_path}, 错误: {e}")

            # 手动删除文件记录 (或者依赖 volume 的 cascade)
            db.delete(file_obj)

        # C. 删除卷册记录
        db.delete(volume)

    # =========================================================
    # 1.5 清理附件 (数据库记录 + 本地文件 + COS 对象)
    # =========================================================
    attachments = db.query(CaseAttachment).filter(CaseAttachment.case_id == case_id).all()
    for attachment in attachments:
        # 删除本地文件及级联空文件夹
        full_path = os.path.join(CASE_ATTACHMENT_ROOT, attachment.file_path)
        from ..utils.storage_manager import cleanup_local_file
        cleanup_local_file(full_path, CASE_ATTACHMENT_ROOT)
        # Word 文档同步删除 PDF 预览缓存
        if full_path.lower().endswith(('.doc', '.docx')):
            pdf_path = os.path.splitext(full_path)[0] + '.pdf'
            cleanup_local_file(pdf_path, CASE_ATTACHMENT_ROOT)

        # COS 模式：删除 COS 对象及 PDF 预览缓存
        cos_key = getattr(attachment, 'cos_key', None)
        if cos_key and settings.STORAGE_TYPE == "COS":
            try:
                from ..utils.storage_manager import _get_cos_client
                _get_cos_client().delete_object(Bucket=settings.COS_BUCKET, Key=cos_key)
                stem, _ = os.path.splitext(cos_key)
                cache_key = f"preview_cache/{stem}.pdf"
                _get_cos_client().delete_object(Bucket=settings.COS_BUCKET, Key=cache_key)
            except Exception as e:
                print(f"[DeleteCase] COS 附件删除失败: {e}")

        db.delete(attachment)

    # =========================================================
    # 2. 清理财务数据
    # =========================================================
    if case.finance:
        # 由于你在 finance_model.py 中配置了 cascade="all, delete-orphan"
        # 删除 CaseFinance 会自动级联删除 Records, Invoices, Withdrawals
        db.delete(case.finance)

    # =========================================================
    # 3. 案件本身的逻辑删除处理 (原逻辑)
    # =========================================================
    # 在案件号前添加删除标记和唯一ID，防止将来复用案号冲突
    if not case.case_number.startswith("[已删除]"):
        case.case_number = f"[已删除]{case.case_number}-ID{case_id}"

    case.is_deleted = True

    # 提交事务
    # SQLAlchemy 会在一个事务中执行上述所有的 delete 和 update
    db.commit()
    return True


def list_cases_by_lawyer(db: Session, lawyer_id: int) -> List[Case]:
    """
    获取指定律师相关的案件（主办/助理/执行律师/执行助理）
    """
    return cast(
        List[Case],
        cast(object, db.query(Case)
             .options(
            joinedload(Case.main_lawyer),
            joinedload(Case.assistant_lawyer),
            joinedload(Case.assistant_lawyer_2),
            joinedload(Case.execution_lawyer),
            joinedload(Case.execution_assistant),
        )
             .filter(
            Case.is_deleted == False,
            (
                    (Case.main_lawyer_id == lawyer_id)
                    | (Case.assistant_lawyer_id == lawyer_id)
                    | (Case.assistant_lawyer_2_id == lawyer_id)
                    | (Case.execution_lawyer_id == lawyer_id)
                    | (Case.execution_assistant_id == lawyer_id)
            )
        )
             .all()),
    )

def count_main_cases(db: Session, lawyer_id: int, year: Optional[int] = None) -> int:
    """统计主办案件数量"""
    query = db.query(Case).filter(Case.main_lawyer_id == lawyer_id, Case.is_deleted == False)
    if year:
        query = query.filter(Case.case_number.like(f"%({year})%"))
    return query.count()


def sum_main_case_income(db: Session, lawyer_id: int, year: Optional[int] = None) -> float:
    """统计主办案件总收费"""
    query = db.query(func.sum(Case.case_income)).filter(Case.main_lawyer_id == lawyer_id, Case.is_deleted == False)
    if year:
        query = query.filter(Case.case_number.like(f"%({year})%"))
    result = query.first()
    return result[0] or 0


def count_cases_by_category(db: Session, lawyer_id: int, year: Optional[int] = None) -> dict:
    """按案件类型统计数量"""
    query = db.query(Case.case_category, func.count(Case.case_id)). \
        filter(Case.main_lawyer_id == lawyer_id, Case.is_deleted == False)

    if year:
        query = query.filter(Case.case_number.like(f"%({year})%"))

    categories = query.group_by(Case.case_category).all()
    return {category: count for category, count in categories}


# 拆分字符串工具
def split_with_separators(s: str, separators: list) -> list:
    """按多个分隔符拆分字符串"""
    import re
    # 构建正则表达式：匹配任何分隔符
    separator_pattern = '|'.join(re.escape(sep) for sep in separators)
    return re.split(separator_pattern, s)


# 事件提醒功能
def get_upcoming_events(
        db: Session,
        user_id: int,
        days: int = 30,
        can_view_all_bank_events: bool = False,
        main_lawyer_id: Optional[int] = None,  # 主办律师筛选
        relation_filter: str = "all",  # 接收前端传来的关系筛选参数
        skip: int = 0,  # 分页
        limit: int = 50  # 分页
) -> dict:
    from datetime import date, timedelta
    from sqlalchemy.orm import joinedload
    from sqlalchemy import or_, and_, between

    today = date.today()
    # 只有 days > 0 时才需要计算目标日期
    if days > 0:
        target_date = today + timedelta(days=days)

    events = []

    # ================= 1. 获取系统提取的案件节点 =================
    # 基础案件范围条件（我参与的）
    user_involved_cond = or_(
        Case.main_lawyer_id == user_id,
        Case.assistant_lawyer_id == user_id,
        Case.assistant_lawyer_2_id == user_id,
        Case.execution_lawyer_id == user_id,
        Case.execution_assistant_id == user_id
    )

    # 安全的“非我参与”条件，防止 SQL NULL 值穿透导致数据丢失
    not_mine_cond = and_(
        or_(Case.main_lawyer_id != user_id, Case.main_lawyer_id.is_(None)),
        or_(Case.assistant_lawyer_id != user_id, Case.assistant_lawyer_id.is_(None)),
        or_(Case.assistant_lawyer_2_id != user_id, Case.assistant_lawyer_2_id.is_(None)),
        or_(Case.execution_lawyer_id != user_id, Case.execution_lawyer_id.is_(None)),
        or_(Case.execution_assistant_id != user_id, Case.execution_assistant_id.is_(None))
    )

    # 严格区分关系过滤
    if relation_filter == "mine":
        final_case_cond = user_involved_cond
    elif relation_filter == "others":
        if can_view_all_bank_events:
            # 修复：使用 not_mine_cond 替代原本的 ~user_involved_cond
            final_case_cond = and_(Case.case_category == "银行案件", not_mine_cond)
        else:
            # 如果没有全行查看权限，不可能看到他人的案件
            final_case_cond = Case.case_id == -1
    else:  # "all"
        if can_view_all_bank_events:
            final_case_cond = or_(user_involved_cond, Case.case_category == "银行案件")
        else:
            final_case_cond = user_involved_cond

    # 构建数据库层面的日期过滤条件
    if days > 0:
        date_conditions = [
            between(Case.hearing_date, today, target_date),
            between(Case.preservation_end, today, target_date),
            between(Case.mediation_due_date, today, target_date),
            between(Case.execution_due_date, today, target_date),
            between(Case.payment_due_date, today, target_date),
            between(Case.advisory_due_date, today, target_date),
        ]
        if can_view_all_bank_events:
            date_conditions.extend([
                between(BankCase.statute_of_limitations, today, target_date),
                between(BankCase.execution_recovery_date, today, target_date),
                between(BankCase.guarantee_due_date, today, target_date),
                between(BankCase.freeze_end_date, today, target_date),
                between(BankCase.seizure_end_date, today, target_date),
            ])
    else:
        date_conditions = [
            Case.hearing_date >= today,
            Case.preservation_end >= today,
            Case.mediation_due_date >= today,
            Case.execution_due_date >= today,
            Case.payment_due_date >= today,
            Case.advisory_due_date >= today,
        ]
        if can_view_all_bank_events:
            date_conditions.extend([
                BankCase.statute_of_limitations >= today,
                BankCase.execution_recovery_date >= today,
                BankCase.guarantee_due_date >= today,
                BankCase.freeze_end_date >= today,
                BankCase.seizure_end_date >= today,
            ])

    # 组装基础查询
    case_query = db.query(Case).outerjoin(BankCase).options(
        joinedload(Case.bank_case_details),
        joinedload(Case.parties)
    ).filter(
        Case.is_deleted == False,
        final_case_cond,
        or_(*date_conditions)
    )

    # 叠加前端传来的主办律师筛选
    if main_lawyer_id:
        case_query = case_query.filter(Case.main_lawyer_id == main_lawyer_id)

    cases = case_query.all()

    # 内存中组装 Event 列表
    for case in cases:
        # 定义需要检查的字段映射
        check_points = [
            ("开庭", case.hearing_date),
            ("保全到期", case.preservation_end),
            ("调解到期", case.mediation_due_date),
            ("执行到期", case.execution_due_date),
            ("付款到期", case.payment_due_date),
            ("顾问到期", case.advisory_due_date),
        ]

        # ========== 将银行案件的诉讼时效加入检查点 ==========
        if case.case_category == "银行案件" and case.bank_case_details:
            check_points.append(("诉讼时效到期", case.bank_case_details.statute_of_limitations))
            check_points.append(("恢复执行时间", case.bank_case_details.execution_recovery_date))
            check_points.append(("保证到期", case.bank_case_details.guarantee_due_date))
            check_points.append(("冻结到期", case.bank_case_details.freeze_end_date))
            check_points.append(("查封到期", case.bank_case_details.seizure_end_date))

        # 动态获取当事人列表中的委托人名称
        clients = [p.name for p in case.parties if p.party_type and '委托' in p.party_type and p.name]
        real_client_name = "、".join(clients) if clients else ""

        is_mine = (
            case.main_lawyer_id == user_id or
            case.assistant_lawyer_id == user_id or
            case.assistant_lawyer_2_id == user_id or
            case.execution_lawyer_id == user_id or
            case.execution_assistant_id == user_id
        )

        for event_type, event_date in check_points:
            if event_date:
                # 检查日期是否符合条件：
                # 1. 如果 days == 0，表示查询全部未来的待办（只限制大于等于今天即可）
                # 2. 如果 days > 0，表示查询 [今天, 目标日期] 范围内的待办
                if (days == 0 and event_date >= today) or (days > 0 and today <= event_date <= target_date):
                    events.append({
                        "case_id": case.case_id,
                        "case_number": case.case_number,
                        "client_name": real_client_name,
                        "event_type": event_type,
                        "event_date": event_date,
                        "days_remaining": (event_date - today).days,
                        "source": "case",
                        "is_mine": is_mine,
                    })

    # ================= 2. 自动清理：删除已过期的自定义日程 =================
    db.query(UserSchedule).filter(UserSchedule.event_date < today).delete(
        synchronize_session=False
    )
    db.commit()

    # ================= 3. 获取用户自定义的日程 =================
    # 只有在筛选“全部”或“我的案件”时，才抓取当前用户的自定义日程
    if relation_filter in ["all", "mine"]:
        schedules_query = db.query(UserSchedule).outerjoin(Case, UserSchedule.related_case_id == Case.case_id)

        schedule_self_cond = UserSchedule.user_id == user_id
        schedule_related_cond = and_(
            UserSchedule.related_case_id != None,
            Case.is_deleted == False,
            user_involved_cond
        )

        if can_view_all_bank_events:
            schedule_bank_cond = and_(
                UserSchedule.related_case_id != None,
                Case.is_deleted == False,
                Case.case_category == "银行案件"
            )
            final_schedule_cond = or_(schedule_self_cond, schedule_related_cond, schedule_bank_cond)
        else:
            final_schedule_cond = or_(schedule_self_cond, schedule_related_cond)

        schedules_query = schedules_query.filter(final_schedule_cond)

        if main_lawyer_id:
            schedules_query = schedules_query.filter(Case.main_lawyer_id == main_lawyer_id)

        if days > 0:
            schedules_query = schedules_query.filter(between(UserSchedule.event_date, today, target_date))
        else:
            schedules_query = schedules_query.filter(UserSchedule.event_date >= today)

        # 预加载可能关联的案件
        custom_schedules = schedules_query.options(
            joinedload(UserSchedule.related_case).joinedload(Case.parties)
        ).all()

        for sched in custom_schedules:
            c_num = sched.related_case.case_number if sched.related_case else None
            c_client = None
            if sched.related_case:
                clients = [p.name for p in sched.related_case.parties if
                           p.party_type and '委托' in p.party_type and p.name]
                c_client = "、".join(clients) if clients else ""

            # 判定“业务归属”：如果有关联案件，当前用户是不是这个案件的参与律师？
            is_involved = False
            if sched.related_case:
                is_involved = (
                        sched.related_case.main_lawyer_id == user_id or
                        sched.related_case.assistant_lawyer_id == user_id or
                        sched.related_case.assistant_lawyer_2_id == user_id or
                        sched.related_case.execution_lawyer_id == user_id or
                        sched.related_case.execution_assistant_id == user_id
                )
            else:
                is_involved = True  # 没有关联案件

            events.append({
                "case_id": sched.related_case_id,
                "case_number": c_num,
                "client_name": c_client,
                "event_type": sched.title,  # 自定义标题作为事件类型
                "event_date": sched.event_date,
                "days_remaining": (sched.event_date - today).days,
                "source": "custom", # 标记为自定义
                "is_mine": is_involved,               # 业务归属判断（我的业务 vs 他人业务）
                "is_creator": sched.user_id == user_id, # 判断是不是当前用户亲自创建的
                "schedule_id": sched.id,
                "description": sched.description
            })

    # ================= 4. 统一排序并分页返回 =================
    events.sort(key=lambda x: x['days_remaining'])

    total = len(events)
    # 模拟分页切片
    paginated_events = events[skip: skip + limit]

    return {
        "items": paginated_events,
        "total": total
    }


def export_cases_to_excel(
        db: Session,
        user_id: int,
        role: str,
        query_params: CaseExportQuery,
        can_view_all_bank: bool = False,
) -> BytesIO:
    """
    导出业务数据为Excel文件 (主表看概况 + 子表查详情)
    优化版：采用 WriteOnly 模式，纯 Python 预计算
    """
    # 1. 基础查询与预加载 (保持不变)
    query = db.query(Case).options(
        selectinload(Case.main_lawyer),
        selectinload(Case.assistant_lawyer),
        selectinload(Case.assistant_lawyer_2),
        selectinload(Case.execution_lawyer),
        selectinload(Case.execution_assistant),
        selectinload(Case.reviewer),
        selectinload(Case.parties),
        selectinload(Case.bank_case_details)
    ).filter(Case.is_deleted == False)

    # 权限控制
    if role not in ["admin", "owner"]:
        user_involved_cond = or_(
            Case.main_lawyer_id == user_id,
            Case.assistant_lawyer_id == user_id,
            Case.assistant_lawyer_2_id == user_id,
            Case.execution_lawyer_id == user_id,
            Case.execution_assistant_id == user_id
        )
        if can_view_all_bank:
            query = query.filter(or_(user_involved_cond, Case.case_category == "银行案件"))
        else:
            query = query.filter(user_involved_cond)

    # 2. 动态筛选条件
    if query_params.case_ids:
        # 如果传入了具体的 ID 列表，则直接过滤这些 ID (精准导出选中)
        query = query.filter(Case.case_id.in_(query_params.case_ids))
    else:
        # 只有在没有特定 ID 的情况下，才应用常规的模糊搜索和区间过滤
        if query_params.keyword:
            query = query.filter(
                or_(
                    Case.case_number.like(f"%{query_params.keyword}%"),
                    Case.parties.any(CaseParty.name.like(f"%{query_params.keyword}%"))
                )
            )
        if query_params.case_category:
            query = query.filter(Case.case_category == query_params.case_category)
        if query_params.main_lawyer_id:
            query = query.filter(Case.main_lawyer_id == query_params.main_lawyer_id)
        if query_params.execution_lawyer_id:
            query = query.filter(Case.execution_lawyer_id == query_params.execution_lawyer_id)
        if query_params.client_name:
            query = query.filter(
                Case.parties.any(
                    and_(CaseParty.party_type.like('%委托%'), CaseParty.name.like(f"%{query_params.client_name}%"))
                )
            )
        if query_params.case_status:
            query = query.join(BankCase).filter(BankCase.case_status == query_params.case_status)

        if query_params.start_date or query_params.end_date:
            if query_params.start_date:
                query = query.filter(Case.commission_date >= query_params.start_date)
            if query_params.end_date:
                query = query.filter(Case.commission_date <= query_params.end_date)
        elif query_params.year:
            query = query.filter(Case.case_number.like(f"%({query_params.year})%"))

    cases = query.order_by(Case.created_at.desc()).all()

    # ---------------- 表头定义 (保持不变) ----------------
    base_headers_part1 = ["业务ID", "业务号", "委托日期", "业务类别"]
    party_headers = [
        "委托人", "委托人联系电话", "委托人证件号",
        "原告/申请人/上诉人", "原告/申请人/上诉人联系电话", "原告/申请人/上诉人证件号",
        "被告(人)/被申请人/被上诉人", "被告(人)/被申请人/被上诉人联系电话", "被告(人)/被申请人/被上诉人证件号",
        "第三人", "第三人联系电话", "第三人证件号",
        "借款人", "借款人联系电话", "借款人证件号",
        "担保人", "担保人联系电话", "担保人证件号",
        "其他当事人", "其他当事人联系电话", "其他当事人证件号"
    ]
    base_headers_part2 = [
        "案件来源", "收费方式", "风险比例", "诉讼标的额", "案件收入",
        "付款到期日", "案由", "介入阶段", "代理权限", "审理法院", "侦查机关", "检察院", "二审检察机关", "开庭时间",
        "立案日", "结案时间",
        "案件地点", "案件详情", "主办律师", "助理律师", "第二助理律师", "执行主办律师", "执行助理律师", "审核状态",
        "审核人", "审核意见",
        "是否重大", "是否纸质卷宗", "是否解除", "是否笔录", "是否保全", "保全开始日", "保全终止日",
        "案号", "结案状态", "结案方式", "诉讼费缴费时间", "诉讼费缴费金额", "诉讼费退费时间", "诉讼费退费金额",
        "申请执行日", "调解到期日", "执行到期日", "顾问到期日", "创建时间", "更新时间"
    ]
    bank_specific_headers = [
        "支行名称", "案件状态", "银行要求案件状态", "缺少具体材料", "抵/质押物信息", "抵押物位置", "客户经理",
        "贷款类型", "贷款种类", "贷款账号",
        "贷款本金", "诉讼标的金额(含利息)", "信用卡违约金", "借款日", "到期日", "诉讼时效", "保证到期日", "收案日期",
        "取材料人", "诉前催收情况", "盖章日", "材料提交法院日", "承办法官", "裁判时间", "裁判方式",
        "裁判摘要", "支持律师费金额", "被告支付律师费金额", "是否还清", "是否有二审/再审",
        "执行案号", "执行立案时间", "执行法官", "借款人工作单位", "是否为恢复执行", "收取执行材料时间",
        "执行材料提交法院时间", "执行本金金额", "执行律师费金额", "财产调查情况", "网络查控财产情况",
        "承办人执行方案", "法院执行措施", "查封冻结时间", "冻结开始日期", "冻结截止日期",
        "查封开始日期", "查封截止日期", "拍卖程序", "拍卖变卖成交价",
        "执行和解内容", "执行和解到期日", "执行和解案件履行跟踪情况", "终本时间", "终本原因", "终结执行时间", "恢复执行时间", "还清时间",
        "执行回款总金额", "执行回款来源", "执行和解跟进及回款额", "扣划跟进及回款额", "调解案件履行跟踪情况"
    ]
    party_detail_headers = [
        "关联业务号", "业务类别", "当事人名称", "类型",
        "法定代表人", "身份证号/统一社会信用代码", "联系电话", "联系地址"
    ]

    standard_headers = base_headers_part1 + party_headers + base_headers_part2
    bank_headers = base_headers_part1 + party_headers + base_headers_part2 + bank_specific_headers

    # ---------------- 辅助函数 ----------------
    def format_date(d):
        return d.strftime("%Y-%m-%d") if d else ""

    def format_datetime(dt):
        return dt.strftime("%Y-%m-%d %H:%M:%S") if dt else ""

    def format_bool(b):
        return "是" if b else "否"

    def format_decimal(d):
        return float(d) if d is not None else 0.0

    def format_json(j):
        if not j: return ""
        try:
            return json.dumps(j, ensure_ascii=False)
        except:
            return str(j)

    # ---------------- 纯 Python 数据提取 (极速，无 openpyxl 开销) ----------------
    standard_rows = []
    bank_rows = []
    party_rows = []

    for case in cases:
        clients_name, clients_phone, clients_id = [], [], []
        plaintiffs_name, plaintiffs_phone, plaintiffs_id = [], [], []
        defendants_name, defendants_phone, defendants_id = [], [], []
        thirds_name, thirds_phone, thirds_id = [], [], []
        borrowers_name, borrowers_phone, borrowers_id = [], [], []
        guarantors_name, guarantors_phone, guarantors_id = [], [], []
        others_name, others_phone, others_id = [], [], []

        if case.parties:
            for p in case.parties:
                party_rows.append([
                    case.case_number or "", case.case_category or "", p.name or "",
                    p.party_type or "", p.legal_representative or "", p.id_number or "",
                    p.phone or "", p.address or ""
                ])

                ptype = p.party_type or ""
                pname = p.name or "未知姓名"
                p_phone = p.phone or "-"
                p_id = p.id_number or "-"

                if "委托人" in ptype:
                    clients_name.append(pname)
                    clients_phone.append(p_phone)
                    clients_id.append(p_id)
                elif ptype in ["原告", "申请人", "上诉人"]:
                    plaintiffs_name.append(pname)
                    plaintiffs_phone.append(p_phone)
                    plaintiffs_id.append(p_id)
                elif ptype in ["被告", "被告人", "被申请人", "被上诉人"]:
                    defendants_name.append(pname)
                    defendants_phone.append(p_phone)
                    defendants_id.append(p_id)
                elif ptype == "第三人":
                    thirds_name.append(pname)
                    thirds_phone.append(p_phone)
                    thirds_id.append(p_id)
                elif ptype == "借款人":
                    borrowers_name.append(pname)
                    borrowers_phone.append(p_phone)
                    borrowers_id.append(p_id)
                elif ptype == "担保人":
                    guarantors_name.append(pname)
                    guarantors_phone.append(p_phone)
                    guarantors_id.append(p_id)
                else:
                    others_name.append(f"{pname}({ptype})" if ptype else pname)
                    others_phone.append(p_phone)
                    others_id.append(p_id)

        # 智能拼接函数：既能保证位置对齐，又能清理全空的无效符号
        def smart_join(items):
            # 如果列表是空的，直接返回空
            if not items:
                return ""

            # 1. 检查是否所有人都没有这条信息（全为空、或者全是 "-"、"None"）
            if all(not item or str(item).strip() in ("", "-", "None", "nan") for item in items):
                return ""  # 如果全都没有，直接返回空字符串，单元格彻底空白

            # 2. 如果有人有，有人没有，为了保证和姓名的顺序一一对应，必须保留占位符
            # 这里把原本丑陋的 "-" 替换成 "无"，视觉上更清晰专业
            processed_items = []
            for item in items:
                val = str(item).strip() if item else ""
                if val in ("", "-", "None", "nan"):
                    processed_items.append("-")
                else:
                    processed_items.append(val)

            return "、".join(processed_items)

        # 使用智能拼接函数生成当事人列数据
        party_columns_data = [
            smart_join(clients_name), smart_join(clients_phone), smart_join(clients_id),
            smart_join(plaintiffs_name), smart_join(plaintiffs_phone), smart_join(plaintiffs_id),
            smart_join(defendants_name), smart_join(defendants_phone), smart_join(defendants_id),
            smart_join(thirds_name), smart_join(thirds_phone), smart_join(thirds_id),
            smart_join(borrowers_name), smart_join(borrowers_phone), smart_join(borrowers_id),
            smart_join(guarantors_name), smart_join(guarantors_phone), smart_join(guarantors_id),
            smart_join(others_name), smart_join(others_phone), smart_join(others_id)
        ]

        base_data_part1 = [case.case_id, case.case_number, format_date(case.commission_date), case.case_category]
        base_data_part2 = [
            case.case_source or "", case.fee_method or "", case.risk_ratio or "", case.claim_amount or "", format_decimal(case.case_income),
            format_date(case.payment_due_date), case.cause or "", case.stage or "", case.agency_power or "",
            case.court or "", case.investigative_agency or "", case.procuratorate or "",
            case.second_instance_procuratorate or "", format_date(case.hearing_date), format_date(case.filing_date),
            format_date(case.closing_date), case.location or "", case.details or "",
            case.main_lawyer.real_name if case.main_lawyer else "",
            case.assistant_lawyer.real_name if case.assistant_lawyer else "",
            case.assistant_lawyer_2.real_name if case.assistant_lawyer_2 else "",
            case.execution_lawyer.real_name if case.execution_lawyer else "",
            case.execution_assistant.real_name if case.execution_assistant else "",
            case.review_status, case.reviewer.real_name if case.reviewer else "",
            case.review_comment or "",
            format_bool(case.is_major), format_bool(case.has_paper_file), format_bool(case.is_dismissed),
            format_bool(case.has_record), format_bool(case.has_preservation),
            format_date(case.preservation_start), format_date(case.preservation_end),
            case.case_code or "", case.closing_status or "", case.closing_method or "",
            format_date(case.litigation_fee_payment_date), format_decimal(case.litigation_fee_payment_amount),
            format_date(case.litigation_fee_refund_date), format_decimal(case.litigation_fee_refund_amount),
            format_date(case.execution_application_date), format_date(case.mediation_due_date),
            format_date(case.execution_due_date), format_date(case.advisory_due_date),
            format_datetime(case.created_at), format_datetime(case.updated_at)
        ]

        if case.case_category == "银行案件":
            bank = case.bank_case_details
            bank_specific_data = [
                bank.branch_name if bank else "",
                bank.case_status if bank else "",
                bank.bank_required_case_status if bank else "",
                bank.missing_specific_materials if bank else "",
                bank.collateral_info if bank else "",
                bank.collateral_location if bank else "",
                bank.account_manager if bank else "",
                bank.loan_type if bank else "",
                bank.loan_category if bank else "",
                bank.loan_account if bank else "",
                format_decimal(bank.loan_principal if bank else 0),
                format_decimal(bank.litigation_target_amount if bank else 0),
                format_decimal(bank.credit_card_penalty if bank else 0),
                format_date(bank.loan_date if bank else None),
                format_date(bank.loan_due_date if bank else None),
                format_date(bank.statute_of_limitations if bank else None),
                format_date(bank.guarantee_due_date if bank else None),
                format_date(bank.case_acceptance_date if bank else None),
                bank.material_fetcher if bank else "",
                bank.pre_litigation_collection if bank else "",
                format_date(bank.seal_date if bank else None),
                format_date(bank.material_submission_date if bank else None),
                bank.handling_judge if bank else "",
                format_date(bank.judgment_date if bank else None),
                bank.judgment_method if bank else "",
                bank.judgment_summary if bank else "",
                format_decimal(bank.lawyer_fee_supported if bank else 0),
                format_decimal(bank.defendant_paid_lawyer_fee if bank else 0),
                format_bool(bank.is_settled if bank else False),
                format_bool(bank.has_second_instance_or_retrial if bank else False),
                bank.execution_case_number if bank else "",
                format_date(bank.execution_filing_date if bank else None),
                bank.execution_judge if bank else "",
                bank.borrower_work_unit if bank else "",
                format_bool(bank.is_execution_recovery if bank else False),
                format_date(bank.execution_material_receipt_date if bank else None),
                format_date(bank.execution_material_submission_date if bank else None),
                format_decimal(bank.execution_principal if bank else 0),
                format_decimal(bank.execution_lawyer_fee if bank else 0),
                bank.property_investigation if bank else "",
                bank.network_control_status if bank else "",
                bank.execution_plan if bank else "",
                bank.court_execution_measures if bank else "",
                format_date(bank.seizure_freeze_date if bank else None),
                format_date(bank.freeze_start_date if bank else None),
                format_date(bank.freeze_end_date if bank else None),
                format_date(bank.seizure_start_date if bank else None),
                format_date(bank.seizure_end_date if bank else None),
                bank.auction_status if bank else "",
                format_decimal(bank.auction_deal_price if bank else 0),
                bank.execution_settlement_content if bank else "",
                format_date(bank.execution_settlement_due_date if bank else None),
                bank.execution_settlement_tracking if bank else "",
                format_date(bank.procedure_termination_date if bank else None),
                bank.termination_reason if bank else "",
                format_date(bank.execution_conclusion_date if bank else None),
                format_date(bank.execution_recovery_date if bank else None),
                format_date(bank.payoff_date if bank else None),
                format_decimal(bank.execution_collection_amount if bank else 0),
                bank.collection_source if bank else "",
                format_json(bank.execution_settlement_log if bank else None),
                format_json(bank.deduction_log if bank else None),
                bank.mediation_tracking if bank else ""
            ]
            bank_rows.append(base_data_part1 + party_columns_data + base_data_part2 + bank_specific_data)
        else:
            standard_rows.append(base_data_part1 + party_columns_data + base_data_part2)

    # ---------------- 3. 极速导出：WriteOnlyWorkbook ----------------
    # 【终极优化】：开启只写模式，直接将数据冲刷至文件流，不构建内存DOM树
    wb = Workbook(write_only=True)
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")
    wrap_alignment = Alignment(wrap_text=True, vertical="center")

    def write_sheet_data(sheet_title, headers, data_rows, wrap_start=4, wrap_end=24):
        ws = wb.create_sheet(title=sheet_title)

        # 纯 Python 预计算前 100 行的列宽，直接在写入前设定完毕
        widths = [len(str(h).encode('gbk', 'replace')) for h in headers]
        for row in data_rows[:100]:
            for idx, val in enumerate(row):
                if val is not None:
                    max_l = max((len(l.encode('gbk', 'replace')) for l in str(val).split('\n')), default=0)
                    if max_l > widths[idx]:
                        widths[idx] = max_l

        for idx, width in enumerate(widths):
            col_letter = get_column_letter(idx + 1)
            ws.column_dimensions[col_letter].width = min(width + 2, 100)

        # 写入表头 (通过 WriteOnlyCell 注入样式)
        header_cells = []
        for val in headers:
            cell = WriteOnlyCell(ws, value=val)
            cell.font = header_font
            cell.alignment = header_alignment
            header_cells.append(cell)
        ws.append(header_cells)

        # 写入数据 (通过 WriteOnlyCell 注入当事人列的换行样式)
        for row in data_rows:
            ws.append(row)

    # 根据数据动态生成 Sheet
    need_bank = (query_params.case_category == "银行案件") or (not query_params.case_category)
    need_standard = (query_params.case_category != "银行案件") or (not query_params.case_category)

    if need_standard and standard_rows:
        write_sheet_data("常规业务", standard_headers, standard_rows)
    # 如果没筛选类型但常规列表为空，也保底建一个空Sheet
    elif need_standard and not standard_rows and not need_bank:
        write_sheet_data("常规业务", standard_headers, [])

    if need_bank and bank_rows:
        write_sheet_data("银行案件", bank_headers, bank_rows)
    elif need_bank and not bank_rows and not need_standard:
        write_sheet_data("银行案件", bank_headers, [])

    if party_rows:
        # 当事人明细不需要特殊换行，所以 wrap_start 和 wrap_end 设到越界区间
        write_sheet_data("当事人明细", party_detail_headers, party_rows, wrap_start=999, wrap_end=999)

    # 4. 保存并返回
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer