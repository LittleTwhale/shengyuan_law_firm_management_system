# crud/finance_crud.py
from io import BytesIO
from typing import List, Dict, Any

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from sqlalchemy import case, exists, func, or_, and_, extract
from sqlalchemy.orm import Session, selectinload

from ..models.case import Case, CaseParty
from ..models.finance_model import CaseFinance, FinancialRecord, InvoiceRecord, LawyerWithdrawal
from ..models.user import User
from ..schemas.finance_schema import (
    FinancialRecordCreate,
    InvoiceRecordCreate,
    CaseFinanceUpdate,
    FinanceStatsQuery, LawyerWithdrawalCreate
)


# =================================================================
#  内部工具函数：筛选与权限逻辑复用
# =================================================================

def _apply_filters(query, db: Session, params: FinanceStatsQuery, current_user: User):
    """
    通用筛选器：同时应用于“列表查询”和“统计查询”
    1. 处理权限 (User vs Admin)
    2. 处理筛选 (日期、律师、案由)
    """
    # ---------------- 1. 权限控制 (Row-Level Security) ----------------
    # 这里的逻辑对应：普通用户只能看自己的，管理员/财务看全部
    # 注意：权限判定通常由 API 层传入 flag，或者这里判断 role
    # 假设：admin, owner, 或拥有 finance_manage 权限的人可以看到全部

    can_view_all = False
    if current_user.role in ['admin', 'owner']:
        can_view_all = True
    elif current_user.permissions and current_user.permissions.get("finance_manage"):
        can_view_all = True

    if not can_view_all:
        # 普通用户：只能看到自己是主办、助理、执行主办、执行助理的案件
        query = query.filter(
            or_(
                Case.main_lawyer_id == current_user.id,
                Case.assistant_lawyer_id == current_user.id,
                Case.assistant_lawyer_2_id == current_user.id,
                Case.execution_lawyer_id == current_user.id,
                Case.execution_assistant_id == current_user.id
            )
        )

    # ---------------- 2. 业务筛选条件 ----------------
    # 键词搜索 (案件号 或 委托人)
    if params.keyword:
        search = f"%{params.keyword}%"
        query = query.filter(
            or_(
                Case.case_number.ilike(search),
                Case.parties.any(
                    and_(CaseParty.party_type.like('%委托%'), CaseParty.name.ilike(search))
                )
            )
        )

    # 按案件类别 (e.g. 民事、刑事)
    if params.case_category:
        query = query.filter(Case.case_category == params.case_category)

    # 按主办律师
    if params.lawyer_id:
        query = query.filter(Case.main_lawyer_id == params.lawyer_id)

    # 按年份 (基于案件委托日期 commission_date)
    if params.year:
        query = query.filter(extract('year', Case.commission_date) == params.year)

    # 按日期范围 (基于案件委托日期)
    if params.start_date:
        query = query.filter(Case.commission_date >= params.start_date)
    if params.end_date:
        query = query.filter(Case.commission_date <= params.end_date)

    return query


def _recalculate_finance_summary(db: Session, finance_id: int):
    """
    当流水发生变化时，自动重新计算 CaseFinance 的汇总字段。
    将收入/退费两次查询合并为一条 CASE WHEN 聚合，减少数据库往返。
    """
    finance = db.query(CaseFinance).filter(CaseFinance.id == finance_id).first()
    if not finance:
        return

    # 1. 合并查询：一次 SQL 同时计算累计实收 (income) 和累计退费 (refund)
    record_totals = db.query(
        func.sum(case((FinancialRecord.record_type == 'income', FinancialRecord.amount), else_=0)),
        func.sum(case((FinancialRecord.record_type == 'refund', FinancialRecord.amount), else_=0)),
    ).filter(FinancialRecord.finance_id == finance_id).first()

    income_total = record_totals[0] or 0
    refund_total = record_totals[1] or 0
    finance.total_received_amount = income_total - refund_total
    finance.total_refund_amount = refund_total

    # 2. 计算累计开票
    invoice_total = db.query(func.sum(InvoiceRecord.invoice_amount)).filter(
        InvoiceRecord.finance_id == finance_id
    ).scalar() or 0
    finance.total_invoiced_amount = invoice_total

    # 3. 累计领款
    withdrawal_total = db.query(func.sum(LawyerWithdrawal.amount)).filter(
        LawyerWithdrawal.finance_id == finance_id
    ).scalar() or 0
    finance.total_withdrawal_amount = withdrawal_total

    # 4. 计算余额/欠款逻辑
    # 判定是否为"标准固定收费"模式
    has_fixed_contract = finance.contract_amount and finance.contract_amount > 0
    is_risk_agency = finance.risk_agency_content and len(finance.risk_agency_content.strip()) > 0

    if has_fixed_contract and not is_risk_agency:
        # 标准模式：自动计算，覆盖可能的人工输入
        finance.unpaid_amount = finance.contract_amount - finance.total_received_amount
        finance.uninvoiced_amount = finance.contract_amount - finance.total_invoiced_amount
    else:
        # 风险代理模式 OR 0合同额模式：
        # 系统不自动计算 unpaid_amount 和 uninvoiced_amount
        # 保持数据库中现有的值 (该值由 update_summary 接口手动写入)
        pass

    db.add(finance)
    db.commit()
    db.refresh(finance)


# =================================================================
#  CRUD 操作
# =================================================================

class CRUDFinance:

    # --- 1. 获取案件财务列表 (带分页与筛选) ---
    def get_multi(
            self,
            db: Session,
            current_user: User,
            query_params: FinanceStatsQuery,
            skip: int = 0,
            limit: int = 20
    ) -> (List[CaseFinance], int): # type: ignore

        # 1. 确保数据一致性：补建缺失的 CaseFinance 记录
        #    使用 NOT EXISTS 比 NOT IN 语义更清晰，MySQL 优化器处理更高效
        missing_cases = db.query(Case).filter(
            ~exists().where(CaseFinance.case_id == Case.case_id),
            Case.is_deleted == False
        ).all()
        if missing_cases:
            new_finances = [
                CaseFinance(case_id=case.case_id, contract_amount=case.case_income or 0)
                for case in missing_cases
            ]
            db.add_all(new_finances)
            db.commit()

        # 2. 构建基础查询（不预加载 parties，改为下方手动精准查询委托人，避免加载全部当事人）
        query = db.query(CaseFinance).join(Case, CaseFinance.case_id == Case.case_id)
        query = query.options(
            selectinload(CaseFinance.case).selectinload(Case.main_lawyer),
        )

        # 3. 应用筛选
        query = _apply_filters(query, db, query_params, current_user)

        # 4. 获取总数 (Count)
        total = query.count()

        # 5. 排序与分页
        query = query.order_by(Case.created_at.desc())
        items = query.offset(skip).limit(limit).all()

        # 6. 手动精准查询委托人（仅加载 party_type 含"委托"的当事人，避免加载原告/被告等无关数据）
        if items:
            case_ids = [item.case_id for item in items]
            all_parties = (
                db.query(CaseParty)
                .filter(CaseParty.case_id.in_(case_ids))
                .filter(CaseParty.party_type.like('%委托%'))
                .all()
            )
            # 构建 case_id → 委托人列表 的查找字典
            client_map: Dict[int, List[CaseParty]] = {}
            for party in all_parties:
                client_map.setdefault(party.case_id, []).append(party)
            # 替换每个 case 的 parties 集合为仅含委托人（前端 getClientNames 仍正常工作）
            for item in items:
                if item.case:
                    item.case.parties = client_map.get(item.case_id, [])

        return items, total

    # --- 2. 获取统计数据 (各项之和) ---
    def get_statistics(
            self,
            db: Session,
            current_user: User,
            query_params: FinanceStatsQuery
    ) -> Dict[str, Any]:

        query = db.query(
            func.count(CaseFinance.id).label("count"),
            func.sum(CaseFinance.contract_amount).label("total_contract"),
            func.sum(CaseFinance.total_received_amount).label("total_received"),
            func.sum(CaseFinance.total_invoiced_amount).label("total_invoiced"),
            func.sum(CaseFinance.unpaid_amount).label("total_unpaid"),
            func.sum(CaseFinance.total_refund_amount).label("total_refund"),
        ).join(Case, CaseFinance.case_id == Case.case_id)

        query = _apply_filters(query, db, query_params, current_user)

        result = query.first()

        return {
            "count_records": result.count or 0,
            "total_contract": result.total_contract or 0,
            "total_income": result.total_received or 0,
            "total_invoiced": result.total_invoiced or 0,
            "total_unpaid": result.total_unpaid or 0,
            "total_refund": result.total_refund or 0
        }

    # --- 3. 获取单个案件财务详情 (自动初始化) ---
    def get_by_case_id(self, db: Session, case_id: int) -> CaseFinance:
        finance = db.query(CaseFinance).options(
            selectinload(CaseFinance.case).selectinload(Case.main_lawyer),
        ).filter(CaseFinance.case_id == case_id).first()
        if not finance:
            # 懒加载：如果还没有财务记录，则创建一个空的
            case = db.query(Case).filter(Case.case_id == case_id).first()
            initial_amount = 0
            if case and case.case_income:
                initial_amount = case.case_income

            finance = CaseFinance(case_id=case_id, contract_amount=initial_amount)
            db.add(finance)
            db.commit()
            db.refresh(finance)

        # 手动精准查询委托人（仅加载 party_type 含"委托"的当事人）
        if finance.case:
            client_parties = (
                db.query(CaseParty)
                .filter(CaseParty.case_id == case_id)
                .filter(CaseParty.party_type.like('%委托%'))
                .all()
            )
            finance.case.parties = client_parties

        return finance

    # --- 4. 更新财务总表 (修改合同额/备注) ---
    def update_summary(
            self, db: Session, finance_id: int, obj_in: CaseFinanceUpdate
    ) -> CaseFinance:
        db_obj = db.query(CaseFinance).filter(CaseFinance.id == finance_id).first()
        if not db_obj:
            return None

        update_data = obj_in.model_dump(exclude_unset=True)
        for field, value in update_data.items():
            setattr(db_obj, field, value)

        db.add(db_obj)
        db.commit()

        # 保存后立即触发重算逻辑 (以防用户虽然填了手动值，但条件符合自动计算，需要被纠正回来)
        _recalculate_finance_summary(db, finance_id)

        db.refresh(db_obj)
        return db_obj

    # --- 5. 新增收支流水 ---
    def create_record(
            self, db: Session, obj_in: FinancialRecordCreate, operator_id: int
    ) -> FinancialRecord:

        # 1. 创建记录
        db_obj = FinancialRecord(
            **obj_in.model_dump(),
            operator_id=operator_id
        )
        db.add(db_obj)
        db.commit()

        # 2. 触发汇总计算
        _recalculate_finance_summary(db, obj_in.finance_id)

        db.refresh(db_obj)
        return db_obj

    # --- 6. 新增发票记录 ---
    def create_invoice(
            self, db: Session, obj_in: InvoiceRecordCreate, operator_id: int
    ) -> InvoiceRecord:

        db_obj = InvoiceRecord(
            **obj_in.model_dump(),
            operator_id=operator_id
        )
        db.add(db_obj)
        db.commit()

        # 2. 触发汇总计算
        _recalculate_finance_summary(db, obj_in.finance_id)

        db.refresh(db_obj)
        return db_obj

    # --- 7. 删除流水 (用于纠错) ---
    def delete_record(self, db: Session, record_id: int) -> bool:
        record = db.query(FinancialRecord).filter(FinancialRecord.id == record_id).first()
        if not record:
            return False

        finance_id = record.finance_id
        db.delete(record)
        db.commit()

        # 删除后重算
        _recalculate_finance_summary(db, finance_id)
        return True

    # --- 8. 删除发票 (用于纠错) ----
    def delete_invoice(self, db: Session, invoice_id: int) -> bool:
        record = db.query(InvoiceRecord).filter(InvoiceRecord.id == invoice_id).first()
        if not record:
            return False

        finance_id = record.finance_id
        db.delete(record)
        db.commit()

        # 删除后重算 (更新已开票总额)
        _recalculate_finance_summary(db, finance_id)
        return True

    # ---  9. 新增律师领款记录 ---
    def create_withdrawal(
            self, db: Session, obj_in: LawyerWithdrawalCreate, operator_id: int
    ) -> LawyerWithdrawal:

        # 1. 创建记录
        db_obj = LawyerWithdrawal(
            **obj_in.model_dump(),
            operator_id=operator_id
        )
        db.add(db_obj)
        db.commit()

        # 2. 触发汇总计算 (更新 CaseFinance.total_withdrawal_amount)
        _recalculate_finance_summary(db, obj_in.finance_id)

        db.refresh(db_obj)
        return db_obj

    # ---  10. 删除律师领款记录 (用于纠错) ---
    def delete_withdrawal(self, db: Session, withdrawal_id: int) -> bool:
        record = db.query(LawyerWithdrawal).filter(LawyerWithdrawal.id == withdrawal_id).first()
        if not record:
            return False

        finance_id = record.finance_id
        db.delete(record)
        db.commit()

        # 删除后重算
        _recalculate_finance_summary(db, finance_id)
        return True

    # --- 11. 导出 Excel (复用筛选逻辑) ---
    def _build_excel_workbook(self, sheets: list):
        """
        使用 write_only 模式流式写入 Excel，避免为每个单元格创建 Python 对象。
        数据行不设逐格样式以换取写入速度，仅表头保留蓝色背景样式。

        sheets: [(sheet_name, headers, data_rows), ...]
        """
        from openpyxl.cell import WriteOnlyCell

        wb = Workbook(write_only=True)

        # 仅在 write_only 模式下重用的样式对象
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="4F81BD")
        header_alignment = Alignment(horizontal="center", vertical="center")
        header_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        for idx, (sheet_name, headers, data_rows) in enumerate(sheets):
            # write_only 模式下没有默认 Sheet，全部用 create_sheet 创建
            ws = wb.create_sheet(title=sheet_name)

            # ---- 表头行（带样式，用 WriteOnlyCell） ----
            header_cells = []
            for col_num, header_title in enumerate(headers, 1):
                cell = WriteOnlyCell(ws, value=header_title)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = header_border
                header_cells.append(cell)
            ws.append(header_cells)

            # ---- 数据行（纯列表追加，无 Cell 对象开销） ----
            # 数值列提前格式化为字符串，保留千位分隔和小数
            for row_data in data_rows:
                formatted_row = []
                for col_idx, val in enumerate(row_data):
                    if isinstance(val, float):
                        # 格式化浮点数：千位分隔 + 两位小数
                        formatted_row.append(f"{val:,.2f}")
                    else:
                        formatted_row.append(val if val is not None else "")
                ws.append(formatted_row)

            # ---- 列宽（write_only 模式支持设置 column_dimensions） ----
            for col_num, header_title in enumerate(headers, 1):
                col_letter = chr(64 + col_num) if col_num <= 26 else None
                if col_letter:
                    ws.column_dimensions[col_letter].width = max(15, len(header_title) * 2.5)

        return wb

    def export_excel(
            self,
            db: Session,
            current_user: User,
            query_params: FinanceStatsQuery
    ) -> BytesIO:
        """
        生成全量财务数据 Excel，包含4个Sheet
        - 使用 selectinload 代替 joinedload，避免多个一对多关系 JOIN 时的笛卡尔积膨胀
        - 一次全量查询 + selectinload 独立 IN 查询，总共 ~10 条 SQL，避免分页导致的查询数翻倍
        """
        # 1. 构建筛选查询并预加载所有关联数据
        query = db.query(CaseFinance).join(Case, CaseFinance.case_id == Case.case_id)
        query = _apply_filters(query, db, query_params, current_user)

        # selectinload 对每个一对多关系发一条独立的 SELECT ... WHERE id IN (...)，
        # 完全避免多表 JOIN 的笛卡尔积乘法效应
        # 注意：不 selectinload Case.parties，因为会加载每个案件的所有当事人类型（原告/被告等），
        # 而导出只需"委托人"。改为下方手动精准查询，减少 60%+ 的 party 数据量。
        query = query.options(
            selectinload(CaseFinance.case).selectinload(Case.main_lawyer),
            selectinload(CaseFinance.records).selectinload(FinancialRecord.operator),
            selectinload(CaseFinance.invoices).selectinload(InvoiceRecord.operator),
            selectinload(CaseFinance.withdrawals).selectinload(LawyerWithdrawal.lawyer),
            selectinload(CaseFinance.withdrawals).selectinload(LawyerWithdrawal.operator),
        ).order_by(Case.created_at.desc())

        finance_list = query.all()

        # 手动仅查询"委托"类型的当事人，避免 selectinload 拉取全部当事人
        case_ids = [f.case_id for f in finance_list]
        all_parties = (
            db.query(CaseParty.case_id, CaseParty.name)
            .filter(CaseParty.case_id.in_(case_ids))
            .filter(CaseParty.party_type.like('%委托%'))
            .all()
        )
        # 构建 case_id → 委托人名称列表 的查找字典
        client_map: Dict[int, List[str]] = {}
        for case_id, name in all_parties:
            client_map.setdefault(case_id, []).append(name)

        # 2. 预定义表头
        headers_1 = [
            "案件号", "委托人", "案件类别", "主办律师", "立案日期",
            "合同金额", "风险代理约定", "最终收费金额",
            "累计实收", "累计开票", "累计退费", "累计领款",
            "未开票金额", "未付金额(欠款)", "可用余额",
            "财务备注", "创建时间", "最后更新"
        ]
        headers_2 = [
            "关联案件号", "委托人", "流水ID", "类型", "金额", "发生日期",
            "付款人/收款人", "支付方式", "操作人", "备注", "登记时间"
        ]
        headers_3 = [
            "关联案件号", "委托人", "发票记录ID", "开票金额", "开票日期",
            "发票号码", "发票抬头", "税号", "经办人", "备注", "登记时间"
        ]
        headers_4 = [
            "关联案件号", "委托人", "领款记录ID", "领款律师", "领款金额",
            "领款日期", "操作人", "备注", "登记时间"
        ]

        # 3. 一次遍历构建四个 Sheet 的数据（纯列表，不保留 ORM 引用）
        data_1, data_2, data_3, data_4 = [], [], [], []

        for f in finance_list:
            c = f.case
            main_lawyer_name = c.main_lawyer.real_name if c.main_lawyer else ""
            # 提取委托人名称（从预查询的 client_map 中取，避免 N+1 且仅加载委托类型）
            client_names = "、".join(client_map.get(c.case_id, []))
            c_info = [c.case_number, client_names]

            # ---- Sheet 1: 案件财务总表 ----
            invoiced = float(f.total_invoiced_amount or 0)
            received = float(f.total_received_amount or 0)
            withdrawal = float(f.total_withdrawal_amount or 0)

            if received > 0:
                tax = invoiced * 0.15
                risk_fund = min(invoiced * 0.05, 50000.0)
            else:
                tax = 0.0
                risk_fund = 0.0
            balance = received - withdrawal - tax - risk_fund

            data_1.append([
                c.case_number, client_names, c.case_category, main_lawyer_name,
                c.commission_date.strftime("%Y-%m-%d") if c.commission_date else "-",
                float(f.contract_amount or 0),
                f.risk_agency_content or "",
                float(f.final_contract_amount or 0),
                float(f.total_received_amount or 0),
                float(f.total_invoiced_amount or 0),
                float(f.total_refund_amount or 0),
                float(f.total_withdrawal_amount or 0),
                float(f.uninvoiced_amount or 0),
                float(f.unpaid_amount or 0),
                balance,
                f.remarks or "",
                f.created_at.strftime("%Y-%m-%d") if f.created_at else "",
                f.updated_at.strftime("%Y-%m-%d") if f.updated_at else ""
            ])

            # ---- Sheet 2: 收支流水明细 ----
            for r in f.records:
                op_name = r.operator.real_name if r.operator else ""
                r_type = "收款" if r.record_type == 'income' else "退费"
                data_2.append(c_info + [
                    r.id, r_type, float(r.amount),
                    r.transaction_date.strftime("%Y-%m-%d") if r.transaction_date else "",
                    r.payer or "",
                    r.payment_method or "",
                    op_name,
                    r.remarks or "",
                    r.created_at.strftime("%Y-%m-%d %H:%M") if r.created_at else ""
                ])

            # ---- Sheet 3: 发票开具明细 ----
            for inv in f.invoices:
                op_name = inv.operator.real_name if inv.operator else ""
                data_3.append(c_info + [
                    inv.id, float(inv.invoice_amount),
                    inv.invoice_date.strftime("%Y-%m-%d") if inv.invoice_date else "",
                    inv.invoice_number or "",
                    inv.invoice_title or "",
                    inv.tax_number or "",
                    op_name,
                    inv.remarks or "",
                    inv.created_at.strftime("%Y-%m-%d %H:%M") if inv.created_at else ""
                ])

            # ---- Sheet 4: 律师领款明细 ----
            for w in f.withdrawals:
                lawyer_name = w.lawyer.real_name if w.lawyer else "未知"
                op_name = w.operator.real_name if w.operator else ""
                data_4.append(c_info + [
                    w.id, lawyer_name, float(w.amount),
                    w.withdrawal_date.strftime("%Y-%m-%d") if w.withdrawal_date else "",
                    op_name,
                    w.remarks or "",
                    w.created_at.strftime("%Y-%m-%d %H:%M") if w.created_at else ""
                ])

        # 4. 使用 write_only 模式一次构建所有 Sheet（无 Cell 对象开销）
        wb = self._build_excel_workbook([
            ("案件财务总表", headers_1, data_1),
            ("收支流水明细", headers_2, data_2),
            ("发票开具明细", headers_3, data_3),
            ("律师领款明细", headers_4, data_4),
        ])

        # 5. 保存到内存流
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return output


finance = CRUDFinance()