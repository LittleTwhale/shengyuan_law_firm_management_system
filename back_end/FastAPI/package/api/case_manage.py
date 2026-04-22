# api/case_manage.py
import re
from collections import defaultdict
from datetime import datetime
from decimal import Decimal
from io import BytesIO
from typing import List, Optional
from urllib.parse import quote

from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File
from fastapi.concurrency import run_in_threadpool
from fastapi.responses import StreamingResponse
from openpyxl import load_workbook
from sqlalchemy import or_
from sqlalchemy.exc import SQLAlchemyError
from sqlalchemy.orm import Session, selectinload, joinedload

from .deps import get_current_active_user
from ..crud.case import list_cases_by_user_role, get_case_by_id, count_cases_by_user_role, create_case, update_case, \
    delete_case, list_bank_cases_by_user_role, count_bank_cases_by_user_role, \
    split_with_separators, export_cases_to_excel
from ..crud.user import get_all_lawyers, get_user_id_by_name
from ..database.database import get_db
from ..models.case import Case, CaseParty, BankCase
from ..models.user import User
from ..schemas.case import CaseOut, CasePageOut, CaseSimpleOut, CaseCreate, CaseUpdate, CaseExportQuery
from ..schemas.user import UserOut
from ..utils.keywords_helper import determine_party_side, get_valid_keywords

router = APIRouter(
    prefix="/cases",
    tags=["case_manage"]
)

# 辅助函数：统一提取并格式化委托人名称
def aggregate_client_names(case_obj: Case) -> str:
    """从 CaseParty 中提取委托人名称并拼接"""
    if not case_obj.parties:
        return case_obj.client_name or ""
    clients = [p.name for p in case_obj.parties if p.party_type and '委托' in p.party_type and p.name]
    # 如果CaseParty中有委托人，优先用CaseParty的拼接结果；否则为了兼容性回落到旧字段
    return "、".join(clients) if clients else (case_obj.client_name or "")

# 辅助函数：统一提取并格式化借款人名称
def aggregate_borrower_names(case_obj: Case) -> str:
    """从 CaseParty 中提取借款人名称并拼接"""
    if not case_obj.parties:
        return ""
    borrowers = [p.name for p in case_obj.parties if p.party_type == '借款人' and p.name]
    return "、".join(borrowers)

# 1️⃣ 获取正式生效案件列表（分页可选）
@router.get("/", response_model=CasePageOut)
def get_cases(
    skip: int = 0,
    limit: int = 100,
    keyword: Optional[str] = None,  # 新增搜索关键词参数
    category: Optional[str] = None,  # 新增案件类别参数
    main_lawyer_id: Optional[int] = None,  # 新增主办律师参数
    execution_lawyer_id: Optional[int] = None, # 新增执行主办律师参数
    year: Optional[str] = None,  # 新增年份参数
    sort_field: Optional[str] = "created_at",  # 排序参数
    sort_dir: Optional[str] = "desc",  # 排序方式
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)  # 安全依赖
):
    """
    获取案件列表
    """
    # 提取权限标识
    user_perms = current_user.permissions or {}
    can_view_all_bank = user_perms.get("can_view_all_bank_events", False)

    cases = list_cases_by_user_role(
        db=db,
        user_id=current_user.id,
        role=current_user.role,
        skip=skip,
        limit=limit,
        keyword=keyword,  # 传递给CRUD函数
        category=category,  # 传递给CRUD函数
        main_lawyer_id=main_lawyer_id,
        execution_lawyer_id=execution_lawyer_id,
        year=year,
        sort_field=sort_field,
        sort_dir=sort_dir,
        can_view_all_bank=can_view_all_bank,
    )
    total = count_cases_by_user_role(
        db=db,
        user_id=current_user.id,
        role=current_user.role,
        keyword=keyword,  # 传递给统计函数
        category=category,  # 传递给统计函数
        main_lawyer_id=main_lawyer_id,
        execution_lawyer_id=execution_lawyer_id,
        year=year,
        can_view_all_bank=can_view_all_bank,
    )
    # 拦截转换：用 CaseParty 覆盖 client_name
    cases_simple = []
    for c in cases:
        simple = CaseSimpleOut.model_validate(c)
        simple.client_name = aggregate_client_names(c)
        cases_simple.append(simple)
    return {"items": cases_simple, "total": total}

# 2️⃣ 获取银行案件列表
@router.get("/bank_cases", response_model=CasePageOut)
def get_bank_cases(
    skip: int = 0,
    limit: int = 100,
    keyword: Optional[str] = None,  # 搜索关键词参数
    main_lawyer_id: Optional[int] = None,
    execution_lawyer_id: Optional[int] = None,
    client_name: Optional[str] = None,
    year: Optional[str] = None,
    case_status: Optional[str] = None,
    sort_field: Optional[str] = "created_at",  # 排序参数
    sort_dir: Optional[str] = "desc",  # 排序方式
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)  # 安全依赖
):
    """
    获取银行案件列表
    """
    # 提取权限标识
    user_perms = current_user.permissions or {}
    can_view_all_bank = user_perms.get("can_view_all_bank_events", False)

    cases = list_bank_cases_by_user_role(
        db=db,
        user_id=current_user.id,
        role=current_user.role,
        skip=skip,
        limit=limit,
        keyword=keyword,
        main_lawyer_id=main_lawyer_id,
        execution_lawyer_id=execution_lawyer_id,
        client_name=client_name,
        year=year,
        case_status=case_status,
        sort_field=sort_field,
        sort_dir=sort_dir,
        can_view_all_bank=can_view_all_bank,
    )
    total = count_bank_cases_by_user_role(
        db=db,
        user_id=current_user.id,
        role=current_user.role,
        keyword=keyword,  # 传递给统计函数
        main_lawyer_id=main_lawyer_id,
        execution_lawyer_id=execution_lawyer_id,
        client_name=client_name,
        year=year,
        case_status=case_status,
        can_view_all_bank=can_view_all_bank,
    )
    # 拦截转换：用 CaseParty 覆盖 client_name
    cases_simple = []
    for c in cases:
        simple = CaseSimpleOut.model_validate(c)
        simple.client_name = aggregate_client_names(c)
        simple.borrower_name = aggregate_borrower_names(c)
        if c.bank_case_details:
            simple.case_status = c.bank_case_details.case_status
        cases_simple.append(simple)
    return {"items": cases_simple, "total": total}


# 5️⃣ 导出案件表格
@router.post("/export", response_class=StreamingResponse)
async def export_cases(
        query: CaseExportQuery,
        db: Session = Depends(get_db),
        current_user: User = Depends(get_current_active_user)
):
    """
    根据筛选条件导出案件明细 (支持分Sheet导出普通案件和银行案件)
    """
    # 提取权限标识
    user_perms = current_user.permissions or {}
    can_view_all_bank = user_perms.get("can_view_all_bank_events", False)

    # 将重度 CPU 计算任务扔进 FastAPI 的底层线程池执行，坚决不阻塞主 Event Loop
    excel_io = await run_in_threadpool(export_cases_to_excel, db, current_user.id, current_user.role, query, can_view_all_bank)

    # --- 获取文件真实大小，用于前端进度条 ---
    excel_io.seek(0, 2)  # 将指针移动到文件流末尾
    file_size = excel_io.tell()  # 获取当前字节数（即文件大小）
    excel_io.seek(0)  # 务必将指针移回开头，否则前端下载到的是空文件
    # -----------------------------------------------

    filename = f"业务数据明细_{datetime.now().strftime('%Y%m%d%H%M')}.xlsx"
    encoded_filename = quote(filename)

    return StreamingResponse(
        excel_io,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}",
            # 允许前端读取 Content-Length 响应头
            "Access-Control-Expose-Headers": "Content-Disposition, Content-Length",
            # 明确告知前端文件大小
            "Content-Length": str(file_size)
        }
    )

# 2️⃣ 获取单条案件详情
@router.get("/{case_id}", response_model=CaseOut)
def get_case(case_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_active_user)):
    """
    获取案件详情
    """
    case = get_case_by_id(db=db, case_id=case_id)
    if not case:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="案件不存在")
    return case


# 3️⃣ 普通用户提交案件操作申请（新增/修改/删除）
@router.post("/case_create", response_model=CaseOut, status_code=status.HTTP_201_CREATED)
def create_new_case(case_in: CaseCreate, db: Session = Depends(get_db), current_user: User = Depends(get_current_active_user)):
    """
    创建新案件
    """
    try:
        new_case = create_case(db=db, case_in=case_in)
        return new_case
    except ValueError as e:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))

@router.put("/case_update/{case_id}", response_model=CaseOut)
def update_existing_case(case_id: int, case_in: CaseUpdate, db: Session = Depends(get_db), current_user: User = Depends(get_current_active_user)):
    """
    更新案件
    """
    updated_case = update_case(db=db, case_id=case_id, case_in=case_in)
    if not updated_case:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="案件不存在")
    return updated_case

@router.delete("/case_delete/{case_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_existing_case(case_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_active_user)):
    """
    删除案件（逻辑删除）
    """
    success = delete_case(db=db, case_id=case_id)
    if not success:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="案件不存在")
    return


# 4️⃣ 获取所有律师列表
@router.get("/users/lawyers", response_model=List[UserOut])
def list_lawyers(db: Session = Depends(get_db), current_user: User = Depends(get_current_active_user)):
    return get_all_lawyers(db)


# 利益冲突检测
@router.post("/check_conflict", status_code=status.HTTP_200_OK)
def check_interest_conflict(
        case_data: CaseCreate,
        db: Session = Depends(get_db),
        current_user: User = Depends(get_current_active_user)
):
    """
    基于 CaseParty 表的全维度利益冲突检测 (混合匹配版：确切+模糊)
    (接入两步法 + selectinload 极致优化，并包含银行支行双重校验)
    """
    input_parties = case_data.parties or []

    # ====== 提取新案件的支行名称  ======
    branch_name = None
    if case_data.bank_case_details and case_data.bank_case_details.branch_name:
        branch_name = case_data.bank_case_details.branch_name.strip()
    # ====================================================

    # ---------------------------------------------------------
    # 1. 提取当前案件的委托人 (New Clients)
    # ---------------------------------------------------------
    new_client_names = set()
    for p in input_parties:
        # 使用模糊匹配识别委托人类型（以防录入成"第一委托人"等）
        if p.party_type and "委托" in p.party_type and p.name:
            new_client_names.add(p.name.strip())

    if not new_client_names and case_data.client_name:
        new_client_names.add(case_data.client_name.strip())

    if not new_client_names:
        return {"has_conflict": False, "details": []}

    # ---------------------------------------------------------
    # 2. 确定委托人阵营 & 锁定对手 (New Opponents)
    # ---------------------------------------------------------
    client_side = "A"  # 默认为原告方
    found_side = None
    has_side_b = False
    has_side_a = False

    for p in input_parties:
        p_name = p.name.strip() if p.name else ""
        if not p_name: continue

        # 名字双向包含判断：这个当事人是我们的委托人吗？(解决全称/简称不一致)
        is_our_client = any(
            (client_name in p_name or p_name in client_name)
            for client_name in new_client_names
        )

        current_side = determine_party_side(p.party_type)
        if current_side == "B":
            has_side_b = True
        elif current_side == "A":
            has_side_a = True

        # 如果确认是委托人，记录下他的诉讼阵营
        if is_our_client and found_side is None:
            if current_side in ["A", "B"]:
                found_side = current_side

    # 推断委托人阵营
    if found_side:
        client_side = found_side
    else:
        if has_side_b:
            client_side = "A"
        elif has_side_a:
            client_side = "B"
        else:
            client_side = "A"

    # ---------------------------------------------------------
    # 3. 提取本案对手名字
    # ---------------------------------------------------------
    target_side_to_find = "B" if client_side == "A" else "A"
    new_case_opponents = set()

    for p in input_parties:
        p_name = p.name.strip() if p.name else ""
        if not p_name: continue

        if determine_party_side(p.party_type) == target_side_to_find:
            # 二次拦截：绝不能把（被模糊匹配为）委托人自己当成对手，防止左右互搏
            is_self = any((client_name in p_name or p_name in client_name) for client_name in new_client_names)
            if not is_self:
                new_case_opponents.add(p_name)

    # 兼容旧字段
    if not new_case_opponents and case_data.defendant and client_side == "A":
        separators = ["、", ",", "，", " ", "；", ";"]
        new_case_opponents = set(d.strip() for d in split_with_separators(case_data.defendant, separators) if d.strip())

    precise_conflicts = []
    processed_conflicts = set()

    # =========================================================================
    # 检测 A: 代理冲突 (起诉现有客户)
    # =========================================================================
    valid_opponents = get_valid_keywords(new_case_opponents)

    if valid_opponents:
        like_conditions = [CaseParty.name.like(f"%{opp}%") for opp in valid_opponents]

        # 第一步仅获取极轻量的整数 case_id，避免直接查出整个ORM对象
        matched_case_ids_query = db.query(CaseParty.case_id).join(Case).filter(
            CaseParty.party_type.like('%委托%'),
            or_(*like_conditions),
            Case.is_deleted == False
        ).distinct().all()
        matched_case_ids = [r[0] for r in matched_case_ids_query]

        if matched_case_ids:
            # 用 selectinload 精准预拉取 Case，以及关联的 parties、银行详情和主审律师
            existing_cases = db.query(Case).options(
                selectinload(Case.parties),
                selectinload(Case.bank_case_details),
                joinedload(Case.main_lawyer)
            ).filter(Case.case_id.in_(matched_case_ids)).all()

            for c in existing_cases:
                match_level = None
                matched_db_name = ""
                for p in c.parties:
                    if p.party_type and '委托' in p.party_type:
                        db_name = p.name.strip() if p.name else ""

                        # 1. 优先判断是否为确切匹配（跟用户输入的某个原名一模一样）
                        if db_name in new_case_opponents:
                            match_level = "exact"
                            matched_db_name = db_name
                            break
                        else:
                            # 2. 如果不是确切匹配，再判断是否为模糊匹配（双向包含核心词）
                            for opp in valid_opponents:
                                if opp in db_name or db_name in opp:
                                    # ====== 银行案件支行双重校验 ======
                                    is_bank_case_db = c.case_category == "银行案件"
                                    is_generic_bank_name = any(gb in opp for gb in ["银行", "农商行", "信用社"])

                                    if is_generic_bank_name:
                                        # 1. 如果新案件填了支行，要求对方主体名称也得包含这个支行
                                        if branch_name and branch_name not in db_name:
                                            continue  # 支行对不上，不算冲突，跳过

                                        # 2. 如果历史案件是银行案件且有明确支行，要求新案对手词必须包含该支行
                                        if is_bank_case_db and c.bank_case_details and c.bank_case_details.branch_name:
                                            db_branch = c.bank_case_details.branch_name.strip()
                                            if db_branch not in opp:
                                                continue
                                    # ==============================================
                                    match_level = "fuzzy"
                                    matched_db_name = db_name
                                    break
                        if match_level:
                            break

                # 如果既不是完全相等，也不包含核心词（被 SQL LIKE 误杀的），直接跳过
                if not match_level:
                    continue

                conflict_key = (c.case_id, "agency_conflict")
                if conflict_key in processed_conflicts: continue

                # 根据匹配等级动态生成文案
                prefix_text = "冲突匹配" if match_level == "exact" else "疑似冲突"
                match_reason = f"完全匹配 '{matched_db_name}'" if match_level == "exact" else f"匹配到关键字 '{matched_db_name}'"

                conflict_info = {
                    "case_number": c.case_number,
                    "other_lawyer_name": c.main_lawyer.real_name if c.main_lawyer else "未知",
                    "conflict_type": "利益冲突（起诉现有客户）",
                    "match_level": match_level,  # 动态赋值：'exact' 或 'fuzzy'
                    "role": "委托人",
                    "message": f"{prefix_text}：新案件对手方（{match_reason}）是我所现有案件【{c.case_number}】的委托人/顾问单位。"
                }
                precise_conflicts.append(conflict_info)
                processed_conflicts.add(conflict_key)

    # =========================================================================
    # 检测 B: 自益冲突 (正在起诉该客户)
    # =========================================================================
    valid_new_clients = get_valid_keywords(new_client_names)

    if valid_new_clients:
        client_like_conditions = [CaseParty.name.like(f"%{client}%") for client in valid_new_clients]

        # 第一步：只查ID
        history_case_ids_query = db.query(CaseParty.case_id).join(Case).filter(
            or_(*client_like_conditions),
            Case.is_deleted == False,
        ).distinct().all()
        history_case_ids = [r[0] for r in history_case_ids_query]

        if history_case_ids:
            # 第二步：使用 selectinload 拉取
            history_cases = db.query(Case).options(
                selectinload(Case.parties),
                selectinload(Case.bank_case_details),
                joinedload(Case.main_lawyer)
            ).filter(Case.case_id.in_(history_case_ids)).all()

            for c in history_cases:
                matched_party = None
                match_level = None
                matched_db_name = ""

                for p in c.parties:
                    db_name = p.name.strip() if p.name else ""
                    if not db_name: continue

                    # 1. 确切匹配
                    if db_name in new_client_names:
                        match_level = "exact"
                        matched_party = p
                        matched_db_name = db_name
                        break
                    else:
                        # 2. 模糊匹配
                        for c_kw in valid_new_clients:
                            if c_kw in db_name or db_name in c_kw:
                                # ====== 银行案件支行双重校验 ======
                                is_bank_case_db = c.case_category == "银行案件"
                                is_generic_bank_name = any(gb in c_kw for gb in ["银行", "农商行", "信用社"])

                                if is_generic_bank_name:
                                    if branch_name and branch_name not in db_name:
                                        continue

                                    if is_bank_case_db and c.bank_case_details and c.bank_case_details.branch_name:
                                        db_branch = c.bank_case_details.branch_name.strip()
                                        if db_branch not in c_kw:
                                            continue
                                # ==============================================
                                match_level = "fuzzy"
                                matched_party = p
                                matched_db_name = db_name
                                break
                    if match_level:
                        break

                if not match_level or not matched_party:
                    continue

                # 查询该历史案件的委托人 (优化点：直接用预加载进内存的 parties 过滤，坚决不在循环里写 db.query)
                host_clients = [p for p in c.parties if p.party_type and '委托' in p.party_type]
                host_client_names = {hc.name.strip() for hc in host_clients if hc.name}

                # 模糊判断回头客
                is_returning_client = any(
                    matched_db_name in hc_name or hc_name in matched_db_name
                    for hc_name in host_client_names
                )
                if is_returning_client: continue

                # 判定历史阵营
                host_side = "A"
                has_host_as_defendant = any(
                    (hc_name in p.name or p.name in hc_name) and determine_party_side(p.party_type) == "B"
                    for p in c.parties for hc_name in host_client_names
                )
                if has_host_as_defendant: host_side = "B"

                # 判定目标在新案件中的阵营
                target_role_side = determine_party_side(matched_party.party_type)

                if target_role_side != "Unknown" and host_side != target_role_side:
                    conflict_key = (c.case_id, "self_conflict")
                    if conflict_key in processed_conflicts: continue

                    # 根据匹配等级动态生成文案
                    prefix_text = "冲突匹配" if match_level == "exact" else "疑似冲突"
                    match_reason = f"完全匹配 '{matched_db_name}'" if match_level == "exact" else f"匹配到关键字 '{matched_db_name}'"

                    conflict_info = {
                        "case_number": c.case_number,
                        "other_lawyer_name": c.main_lawyer.real_name if c.main_lawyer else "未知",
                        "conflict_type": "利益冲突（正在起诉该客户）",
                        "match_level": match_level,
                        "role": matched_party.party_type,
                        "message": f"{prefix_text}：本案委托人（{match_reason}）在现有案件【{c.case_number}】中是【{matched_party.party_type}】，处于对立面。"
                    }
                    precise_conflicts.append(conflict_info)
                    processed_conflicts.add(conflict_key)

    if precise_conflicts:
        return {"has_conflict": True, "details": precise_conflicts}

    return {"has_conflict": False, "details": []}


@router.post("/import", status_code=200)
def import_cases_from_excel(file: UploadFile = File(...), db: Session = Depends(get_db),
                            current_user: User = Depends(get_current_active_user)):
    """
    📦 批量导入业务接口 (支持双轨制当事人录入)
    """
    try:
        if not file.filename.endswith(('.xlsx', '.xls')):
            raise HTTPException(status_code=400, detail="仅支持.xlsx和.xls格式的Excel文件")

        wb = load_workbook(filename=BytesIO(file.file.read()), data_only=True)

        # 校验两个必须的 Sheet
        if "业务列表" not in wb.sheetnames or "当事人列表" not in wb.sheetnames:
            raise HTTPException(status_code=400, detail="Excel模板错误：必须包含'业务列表'和'当事人列表'两个工作表")

        ws_cases = wb["业务列表"]
        ws_parties = wb["当事人列表"]
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"无法读取Excel文件：{str(e)}")
    finally:
        file.file.close()

    # ---------------- 1. 解析【当事人列表】(高优先级详细数据) ----------------
    parties_headers = [str(cell.value).strip() if cell.value else "" for cell in ws_parties[1]]
    parties_dict = defaultdict(list)  # 格式: { "案件号": [party1, party2, ...] }

    for row in ws_parties.iter_rows(min_row=2, values_only=True):
        p_row = dict(zip(parties_headers, row))
        link_case_no = str(p_row.get("关联业务号", "")).strip()
        p_type = str(p_row.get("当事人类型", "")).strip()
        p_name = str(p_row.get("姓名/名称", "")).strip()

        # 如果缺少关联号、类型或姓名，则跳过该当事人
        if not link_case_no or not p_type or not p_name:
            continue

        parties_dict[link_case_no].append({
            "party_type": p_type,
            "name": p_name,
            "phone": str(p_row.get("联系电话", "")).strip() or None,
            "id_number": str(p_row.get("身份证号/统一社会信用代码", "")).strip() or None,
            "address": str(p_row.get("联系地址", "")).strip() or None,
            "legal_representative": str(p_row.get("法定代表人", "")).strip() or None,
        })

    # ---------------- 2. 解析【业务列表】 ----------------
    cases_headers = [str(cell.value).strip() if cell.value else "" for cell in ws_cases[1]]

    required_cols = ["业务号", "业务类别"]
    for col in required_cols:
        if col not in cases_headers:
            raise HTTPException(status_code=400, detail=f"'业务列表'工作表缺少必要字段：{col}")

    total_rows = 0
    success_rows = 0
    failed_cases = []

    def parse_date(date_value):
        if not date_value: return None
        if isinstance(date_value, str):
            try:
                return datetime.strptime(date_value, "%Y-%m-%d").date()
            except ValueError:
                return None
        return date_value

    def parse_decimal(value):
        if not value: return 0
        try:
            return Decimal(str(value).replace(",", ""))
        except:
            return 0

    for row in ws_cases.iter_rows(min_row=2, values_only=True):
        # 1. 先组装字典
        row_data = dict(zip(cases_headers, row))

        # 2. 提取业务号并进行严格判空
        raw_case_number = row_data.get("业务号")

        # 3. 检查是否为空行（拦截真正的空值，以及字符串"None"、"NaN"等）
        if not raw_case_number or str(raw_case_number).strip().lower() in ["none", "nan", "null", ""]:
            continue

        case_number = str(raw_case_number).strip()

        total_rows += 1

        try:
            # 获取律师ID (兼容为空的情况)
            lawyer_name = str(row_data.get("主办律师", "")).strip()
            main_lawyer_id = None
            if lawyer_name and lawyer_name not in ("None", "nan", "NaN", "null"):
                main_lawyer_id = get_user_id_by_name(db, lawyer_name)
                if not main_lawyer_id:
                    failed_cases.append(
                        {"case_number": case_number, "reason": f"主办律师不存在：{lawyer_name}"})
                    continue

            # ---------------- 3. 解析银行案件专属字段 ----------------
            case_category = str(row_data.get("业务类别", "")).strip()
            bank_details = None

            if case_category == "银行案件":
                bank_details = {
                    "branch_name": str(row_data.get("支行名称", "")).strip() or None,
                    "case_status": str(row_data.get("案件状态", "")).strip() or None,
                    "bank_required_case_status": str(row_data.get("银行要求案件状态", "")).strip() or None,
                    "collateral_info": str(row_data.get("抵/质押物信息", "")).strip() or None,
                    "collateral_location": str(row_data.get("抵押物位置", "")).strip() or None,
                    "account_manager": str(row_data.get("客户经理", "")).strip() or None,
                    "loan_type": str(row_data.get("贷款类型", "")).strip() or None,
                    "loan_account": str(row_data.get("贷款账号", "")).strip() or None,
                    "loan_principal": parse_decimal(row_data.get("贷款本金")),
                    "litigation_target_amount": parse_decimal(row_data.get("诉讼标的金额(含利息)")),
                    "credit_card_penalty": parse_decimal(row_data.get("信用卡违约金")),
                    "loan_date": parse_date(row_data.get("借款日")),
                    "loan_due_date": parse_date(row_data.get("到期日")),
                    "statute_of_limitations": parse_date(row_data.get("诉讼时效")),
                    "guarantee_due_date": parse_date(row_data.get("保证到期日")),
                    "case_acceptance_date": parse_date(row_data.get("收案日期")),
                    "material_fetcher": str(row_data.get("取材料人", "")).strip() or None,
                    "missing_specific_materials": str(row_data.get("缺少具体材料", "")).strip() or None,
                    "pre_litigation_collection": str(row_data.get("诉前催收情况", "")).strip() or None,
                    "seal_date": parse_date(row_data.get("盖章日")),
                    "material_submission_date": parse_date(row_data.get("材料提交法院日")),
                    "handling_judge": str(row_data.get("承办法官", "")).strip() or None,
                    "judgment_date": parse_date(row_data.get("裁判时间")),
                    "judgment_method": str(row_data.get("裁判方式", "")).strip() or None,
                    "judgment_summary": str(row_data.get("裁判摘要", "")).strip() or None,
                    "lawyer_fee_supported": parse_decimal(row_data.get("支持律师费金额")),
                    "defendant_paid_lawyer_fee": parse_decimal(row_data.get("被告支付律师费金额")),
                    "is_settled": (str(row_data.get("是否还清", "")).strip() == "是"),
                    "has_second_instance_or_retrial": (str(row_data.get("是否有二审/再审", "")).strip() == "是"),
                    "execution_case_number": str(row_data.get("执行案号", "")).strip() or None,
                    "execution_filing_date": parse_date(row_data.get("执行立案时间")),
                    "execution_judge": str(row_data.get("执行法官", "")).strip() or None,
                    "borrower_work_unit": str(row_data.get("借款人工作单位", "")).strip() or None,
                    "is_execution_recovery": (str(row_data.get("是否为恢复执行", "")).strip() == "是"),
                    "execution_material_receipt_date": parse_date(row_data.get("收取执行材料时间")),
                    "execution_material_submission_date": parse_date(row_data.get("执行材料提交法院时间")),
                    "execution_principal": parse_decimal(row_data.get("执行本金金额")),
                    "execution_lawyer_fee": parse_decimal(row_data.get("执行律师费金额")),
                    "property_investigation": str(row_data.get("财产调查情况", "")).strip() or None,
                    "network_control_status": str(row_data.get("网络查控财产情况", "")).strip() or None,
                    "execution_plan": str(row_data.get("承办人执行方案", "")).strip() or None,
                    "court_execution_measures": str(row_data.get("法院执行措施", "")).strip() or None,
                    "seizure_freeze_date": parse_date(row_data.get("查封冻结时间")),
                    "auction_status": str(row_data.get("拍卖程序", "")).strip() or None,
                    "auction_deal_price": parse_decimal(row_data.get("拍卖变卖成交价")),
                    "execution_settlement_content": str(row_data.get("执行和解内容", "")).strip() or None,
                    "execution_settlement_due_date": parse_date(row_data.get("执行和解到期日")),
                    "execution_settlement_tracking": str(row_data.get("执行和解案件履行跟踪情况", "")).strip() or None,
                    "procedure_termination_date": parse_date(row_data.get("终本时间")),
                    "termination_reason": str(row_data.get("终本原因", "")).strip() or None,
                    "execution_conclusion_date": parse_date(row_data.get("终结执行时间")),
                    "execution_recovery_date": parse_date(row_data.get("恢复执行时间")),
                    "payoff_date": parse_date(row_data.get("还清时间")),
                    "execution_collection_amount": parse_decimal(row_data.get("执行回款总金额")),
                    "collection_source": str(row_data.get("执行回款来源", "")).strip() or None,
                    "mediation_tracking": str(row_data.get("调解案件履行跟踪情况", "")).strip() or None,
                }

            # ---------------- 4. 合并双轨制当事人列表 ----------------
            matched_detailed_parties = parties_dict.get(case_number, [])
            merged_parties_map = {}

            # 步骤A：优先将【当事人列表】中的详细信息载入字典，以类型+姓名为唯一键
            for p in matched_detailed_parties:
                unique_key = f"{p['party_type']}_{p['name']}"
                merged_parties_map[unique_key] = p

            # 辅助函数：清理 Excel 中的空值（处理 NaN, None 等）
            def clean_excel_val(val):
                if val is None: return None
                s = str(val).strip()
                if s.lower() in ["none", "nan", "null", ""]: return None
                return s

            # 步骤B：解析处理【业务列表】中的快捷字段 (支持单人独立列 + 多人横杠解析)
            def add_quick_parties(party_type, names_str, id_col_str=None, phone_col_str=None):
                names_str = clean_excel_val(names_str)
                if not names_str:
                    return

                # 1. 先按逗号、分号、顿号、换行符分割出多个人
                entries = [n.strip() for n in re.split(r'[;；,，、\n]+', names_str) if n.strip()]

                # 判断当前单元格是否只有一个人
                is_single_person = (len(entries) == 1)

                # 获取独立列的数据
                explicit_id = clean_excel_val(id_col_str)
                explicit_phone = clean_excel_val(phone_col_str)

                for entry in entries:
                    # 2. 对单个人员字符串再次分割，识别 "姓名-身份证号" 格式
                    parts = re.split(r'[-—－_]+', entry, maxsplit=1)

                    name = parts[0].strip()
                    parsed_id = parts[1].strip() if len(parts) > 1 else None

                    # 3. 决定最终要写入的 身份证 和 电话
                    final_id = parsed_id  # 默认使用横杠解析出来的身份证
                    final_phone = None

                    # 如果只有一个人，且独立列填了数据，则优先使用独立列的数据覆盖
                    if is_single_person:
                        if explicit_id:
                            final_id = explicit_id
                        if explicit_phone:
                            final_phone = explicit_phone

                    unique_key = f"{party_type}_{name}"

                    # 4. 去重逻辑：只有在子表没填这个人的时候，才把简略信息加进去
                    if unique_key not in merged_parties_map:
                        merged_parties_map[unique_key] = {
                            "party_type": party_type,
                            "name": name,
                            "phone": final_phone,  # 存入电话
                            "id_number": final_id,  # 存入身份证号
                            "address": None,
                            "legal_representative": None
                        }

            # 载入各类快捷当事人，依次传入：类型、姓名列、身份证列、电话列
            add_quick_parties("原告",
                              row_data.get("快捷原告"),
                              row_data.get("快捷原告身份证号"),
                              row_data.get("快捷原告电话"))

            add_quick_parties("被告",
                              row_data.get("快捷被告"),
                              row_data.get("快捷被告身份证号"),
                              row_data.get("快捷被告电话"))

            add_quick_parties("委托人",
                              row_data.get("快捷委托人"),
                              row_data.get("快捷委托人身份证号"),
                              row_data.get("快捷委托人电话"))

            # 借款人,担保人和第三人也同理应用
            add_quick_parties("借款人",
                              row_data.get("快捷借款人"),
                              row_data.get("快捷借款人身份证号"),
                              row_data.get("快捷借款人电话"))

            add_quick_parties("担保人",
                              row_data.get("快捷担保人"),
                              row_data.get("快捷担保人身份证号"),
                              row_data.get("快捷担保人电话"))

            add_quick_parties("第三人",
                              row_data.get("快捷第三人"),
                              row_data.get("快捷第三人身份证号"),
                              row_data.get("快捷第三人电话"))

            # 生成最终无重复的当事人List
            final_parties = list(merged_parties_map.values())

            # ---------------- 5. 组装 CaseCreate ----------------
            new_case = CaseCreate(
                case_number=case_number,
                commission_date=parse_date(row_data.get("委托日期")),
                case_category=case_category,
                parties=final_parties,  # 🌟 将合并去重后的当事人列表传入
                bank_case_details=bank_details,

                case_source=str(row_data.get("案件来源", "")).strip() or None,
                fee_method=str(row_data.get("收费方式", "")).strip() or None,
                risk_ratio=str(row_data.get("风险比例", "")).strip() or None,
                case_income=parse_decimal(row_data.get("案件收入")),
                payment_due_date=parse_date(row_data.get("付款到期日")),
                cause=str(row_data.get("案由", "")).strip() or None,
                stage=str(row_data.get("介入阶段", "")).strip() or None,
                agency_power=str(row_data.get("代理权限", "")).strip() or None,
                court=str(row_data.get("审理法院", "")).strip() or None,
                investigative_agency=str(row_data.get("侦查机关", "")).strip() or None,
                procuratorate=str(row_data.get("检察院", "")).strip() or None,
                second_instance_procuratorate=str(row_data.get("二审检察机关", "")).strip() or None,
                hearing_date=parse_date(row_data.get("开庭时间")),
                filing_date=parse_date(row_data.get("立案日")),
                closing_date=parse_date(row_data.get("结案时间")),
                location=str(row_data.get("案件地点", "")).strip() or None,
                details=str(row_data.get("案件详情", "")).strip() or None,

                main_lawyer_id=main_lawyer_id,
                assistant_lawyer_id=get_user_id_by_name(db, row_data.get("助理律师")),
                assistant_lawyer_2_id=get_user_id_by_name(db, row_data.get("第二助理律师")),
                execution_lawyer_id=get_user_id_by_name(db, row_data.get("执行主办律师")),
                execution_assistant_id=get_user_id_by_name(db, row_data.get("执行助理律师")),

                is_major=(str(row_data.get("是否重大", "")).strip() == "是"),
                has_paper_file=(str(row_data.get("是否纸质卷宗", "")).strip() == "是"),
                is_dismissed=(str(row_data.get("是否解除", "")).strip() == "是"),
                has_record=(str(row_data.get("是否笔录", "")).strip() == "是"),
                has_preservation=(str(row_data.get("是否保全", "")).strip() == "是"),
                preservation_start=parse_date(row_data.get("保全开始日")),
                preservation_end=parse_date(row_data.get("保全终止日")),

                case_code=str(row_data.get("案号", "")).strip() or None,
                closing_status=str(row_data.get("结案状态", "")).strip() or None,
                closing_method=str(row_data.get("结案方式", "")).strip() or None,

                litigation_fee_payment_date=parse_date(row_data.get("诉讼费缴费时间")),
                litigation_fee_payment_amount=parse_decimal(row_data.get("诉讼费缴费金额")),
                litigation_fee_refund_date=parse_date(row_data.get("诉讼费退费时间")),
                litigation_fee_refund_amount=parse_decimal(row_data.get("诉讼费退费金额")),

                execution_application_date=parse_date(row_data.get("申请执行日")),
                mediation_due_date=parse_date(row_data.get("调解到期日")),
                execution_due_date=parse_date(row_data.get("执行到期日")),
                advisory_due_date=parse_date(row_data.get("顾问到期日"))
            )

            # ---------------- 6. 写入数据库 ----------------
            create_case(db=db, case_in=new_case)
            db.commit()
            success_rows += 1

        except SQLAlchemyError as e:
            db.rollback()
            failed_cases.append({"case_number": case_number, "reason": f"数据库错误：{str(e)}"})
        except Exception as e:
            failed_cases.append({"case_number": case_number, "reason": f"数据处理错误：{str(e)}"})

    return {
        "total_cases": total_rows,
        "imported_cases": success_rows,
        "failed_cases": failed_cases
    }


@router.post("/batch_sync_excel")
async def batch_sync_excel(
        file: UploadFile = File(...),
        db: Session = Depends(get_db),
        current_user: User = Depends(get_current_active_user)
):
    """
    高级批量同步接口：支持通过 Excel 覆盖更新现有案件
    逻辑：若 ID 存在则更新，若 ID 不存在则新增
    """
    try:
        contents = await file.read()
        wb = load_workbook(filename=BytesIO(contents), data_only=True)
        # 默认处理第一个 Sheet (常规业务或银行案件)
        ws = wb.active
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"文件读取失败: {str(e)}")

    headers = [str(cell.value).strip() for cell in ws[1]]

    def parse_date(val):
        if not val or str(val).lower() in ["none", "nan", ""]: return None
        if isinstance(val, datetime): return val.date()
        try:
            return datetime.strptime(str(val), "%Y-%m-%d").date()
        except:
            return None

    def parse_decimal(val):
        if val is None or str(val).lower() in ["none", "nan", ""]: return 0
        try:
            return Decimal(str(val).replace(",", ""))
        except:
            return 0

    def parse_bool(val):
        return str(val).strip() == "是"

    success_count = 0
    update_count = 0
    errors = []

    # 遍历数据行
    for row_idx, row_values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        row_data = dict(zip(headers, row_values))
        case_id = row_data.get("业务ID")
        case_number = str(row_data.get("业务号", "")).strip()

        if not case_number or case_number.lower() in ["none", "nan", ""]:
            continue

        try:
            # 1. 查找现有案件
            existing_case = None
            if case_id:
                existing_case = db.query(Case).filter(Case.case_id == case_id).first()

            # 2. 准备基础字段数据 (全量字段映射)
            case_update_dict = {
                "commission_date": parse_date(row_data.get("委托日期")),
                "case_category": row_data.get("业务类别"),
                "case_source": row_data.get("案件来源"),
                "fee_method": row_data.get("收费方式"),
                "risk_ratio": row_data.get("风险比例"),
                "case_income": parse_decimal(row_data.get("案件收入")),
                "payment_due_date": parse_date(row_data.get("付款到期日")),
                "cause": row_data.get("案由"),
                "stage": row_data.get("介入阶段"),
                "agency_power": row_data.get("代理权限"),
                "court": row_data.get("审理法院"),
                "investigative_agency": row_data.get("侦查机关"),
                "procuratorate": row_data.get("检察院"),
                "second_instance_procuratorate": row_data.get("二审检察机关"),
                "hearing_date": parse_date(row_data.get("开庭时间")),
                "filing_date": parse_date(row_data.get("立案日")),
                "closing_date": parse_date(row_data.get("结案时间")),
                "location": row_data.get("案件地点"),
                "details": row_data.get("案件详情"),
                "is_major": parse_bool(row_data.get("是否重大")),
                "has_paper_file": parse_bool(row_data.get("是否纸质卷宗")),
                "is_dismissed": parse_bool(row_data.get("是否解除")),
                "has_record": parse_bool(row_data.get("是否笔录")),
                "has_preservation": parse_bool(row_data.get("是否保全")),
                "preservation_start": parse_date(row_data.get("保全开始日")),
                "preservation_end": parse_date(row_data.get("保全终止日")),
                "case_code": row_data.get("案号"),
                "closing_status": row_data.get("结案状态"),
                "closing_method": row_data.get("结案方式"),
                "litigation_fee_payment_date": parse_date(row_data.get("诉讼费缴费时间")),
                "litigation_fee_payment_amount": parse_decimal(row_data.get("诉讼费缴费金额")),
                "litigation_fee_refund_date": parse_date(row_data.get("诉讼费退费时间")),
                "litigation_fee_refund_amount": parse_decimal(row_data.get("诉讼费退费金额")),
                "application_execution_date": parse_date(row_data.get("申请执行日")),
                "mediation_due_date": parse_date(row_data.get("调解到期日")),
                "execution_due_date": parse_date(row_data.get("执行到期日")),
                "advisory_due_date": parse_date(row_data.get("顾问到期日")),
            }

            # 3. 处理银行案件专属字段
            bank_data = None
            if row_data.get("业务类别") == "银行案件":
                bank_data = {
                    "branch_name": row_data.get("支行名称"),
                    "case_status": row_data.get("案件状态"),
                    "bank_required_case_status": row_data.get("银行要求案件状态"),
                    "collateral_info": row_data.get("抵/质押物信息"),
                    "collateral_location": row_data.get("抵押物位置"),
                    "account_manager": row_data.get("客户经理"),
                    "loan_type": row_data.get("贷款类型"),
                    "loan_account": row_data.get("贷款账号"),
                    "loan_principal": parse_decimal(row_data.get("贷款本金")),
                    "litigation_target_amount": parse_decimal(row_data.get("诉讼标的金额(含利息)")),
                    "credit_card_penalty": parse_decimal(row_data.get("信用卡违约金")),
                    "loan_date": parse_date(row_data.get("借款日")),
                    "loan_due_date": parse_date(row_data.get("到期日")),
                    "statute_of_limitations": parse_date(row_data.get("诉讼时效")),
                    "guarantee_due_date": parse_date(row_data.get("保证到期日")),
                    "case_acceptance_date": parse_date(row_data.get("收案日期")),
                    "material_fetcher": row_data.get("取材料人"),
                    "pre_litigation_collection": row_data.get("诉前催收情况"),
                    "seal_date": parse_date(row_data.get("盖章日")),
                    "material_submission_date": parse_date(row_data.get("材料提交法院日")),
                    "handling_judge": row_data.get("承办法官"),
                    "judgment_date": parse_date(row_data.get("裁判时间")),
                    "judgment_method": row_data.get("裁判方式"),
                    "judgment_summary": row_data.get("裁判摘要"),
                    "lawyer_fee_supported": parse_decimal(row_data.get("支持律师费金额")),
                    "defendant_paid_lawyer_fee": parse_decimal(row_data.get("被告支付律师费金额")),
                    "is_settled": parse_bool(row_data.get("是否还清")),
                    "has_second_instance_or_retrial": parse_bool(row_data.get("是否有二审/再审")),
                    "execution_case_number": row_data.get("执行案号"),
                    "execution_filing_date": parse_date(row_data.get("执行立案时间")),
                    "execution_judge": row_data.get("执行法官"),
                    "borrower_work_unit": row_data.get("借款人工作单位"),
                    "is_execution_recovery": parse_bool(row_data.get("是否为恢复执行")),
                    "execution_material_receipt_date": parse_date(row_data.get("收取执行材料时间")),
                    "execution_material_submission_date": parse_date(row_data.get("执行材料提交法院时间")),
                    "execution_principal": parse_decimal(row_data.get("执行本金金额")),
                    "execution_lawyer_fee": parse_decimal(row_data.get("执行律师费金额")),
                    "property_investigation": row_data.get("财产调查情况"),
                    "network_control_status": row_data.get("网络查控财产情况"),
                    "execution_plan": row_data.get("承办人执行方案"),
                    "court_execution_measures": row_data.get("法院执行措施"),
                    "seizure_freeze_date": parse_date(row_data.get("查封冻结时间")),
                    "auction_status": row_data.get("拍卖程序"),
                    "auction_deal_price": parse_decimal(row_data.get("拍卖变卖成交价")),
                    "execution_settlement_content": row_data.get("执行和解内容"),
                    "execution_settlement_due_date": parse_date(row_data.get("执行和解到期日")),
                    "execution_settlement_tracking": str(row_data.get("执行和解案件履行跟踪情况", "")).strip() or None,
                    "procedure_termination_date": parse_date(row_data.get("终本时间")),
                    "termination_reason": row_data.get("终本原因"),
                    "execution_conclusion_date": parse_date(row_data.get("终结执行时间")),
                    "execution_recovery_date": parse_date(row_data.get("恢复执行时间")),
                    "payoff_date": parse_date(row_data.get("还清时间")),
                    "execution_collection_amount": parse_decimal(row_data.get("执行回款总金额")),
                    "collection_source": row_data.get("执行回款来源"),
                    "mediation_tracking": row_data.get("调解案件履行跟踪情况"),
                }

            # 4. 执行更新或新增
            if existing_case:
                # 更新 Case 表
                for key, value in case_update_dict.items():
                    setattr(existing_case, key, value)

                # 更新或创建 BankCase 表
                if bank_data:
                    if existing_case.bank_case_details:
                        for key, value in bank_data.items():
                            setattr(existing_case.bank_case_details, key, value)
                    else:
                        new_bank = BankCase(case_id=existing_case.case_id, **bank_data)
                        db.add(new_bank)
                update_count += 1
            else:
                # 新增逻辑 (注意：此处暂不处理当事人，建议当事人仍通过专项 Sheet 导入)
                new_case = Case(case_number=case_number, **case_update_dict)
                db.add(new_case)
                db.flush()
                if bank_data:
                    new_bank = BankCase(case_id=new_case.case_id, **bank_data)
                    db.add(new_bank)
                success_count += 1

            db.commit()

        except Exception as e:
            db.rollback()
            errors.append(f"第{row_idx}行({case_number})处理失败: {str(e)}")

    return {
        "summary": f"同步完成。更新: {update_count}, 新增: {success_count}, 失败: {len(errors)}",
        "errors": errors
    }